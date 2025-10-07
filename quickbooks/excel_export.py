from __future__ import annotations

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .csv_reader import extract_sales_orders_from_csv
from .dataframe_utils import optimize_dataframe_dtypes
from .odoo_enrichment import (
    get_formula_info_from_bom,
    get_odoo_sales_order_line_items,
)


def ensure_excel_datetime_compatibility(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure all datetime-like data is tz-naive and kept as true datetimes for Excel.

    - Converts tz-aware datetimes to tz-naive (keeps wall time).
    - Parses object columns that look like dates into datetime64[ns].
    - Converts individual datetime objects to strings.
    - Leaves non-date columns untouched.
    """
    df_copy = df.copy()
    print(
        f"   🔍 Ensuring Excel datetime compatibility for {len(df_copy.columns)} columns..."
    )

    for col in df_copy.columns:
        series = df_copy[col]

        # Skip empty columns
        if series.empty or series.isna().all():
            continue

        # Case 1: Already a datetime dtype
        if pd.api.types.is_datetime64_any_dtype(series):
            print(f"   📅 Found datetime column: {col} (dtype: {series.dtype})")
            # Handle tz-aware via pandas >= 1.1 behavior (DatetimeTZDtype)
            if hasattr(series.dtype, "tz") and series.dtype.tz is not None:
                try:
                    df_copy[col] = series.dt.tz_localize(None)
                    print(f"   🕐 Removed timezone from column: {col}")
                except Exception as e:
                    print(f"   ⚠️  Could not drop tz for {col}: {e}")
            continue

        # Case 2: Object column that might contain datetime objects or datetime-like strings
        if series.dtype == "object":
            # Check for individual datetime objects in the series
            has_datetime_objects = any(
                isinstance(val, datetime) for val in series.dropna()
            )
            
            if has_datetime_objects:
                print(f"   📅 Found datetime objects in column: {col}")
                # Convert datetime objects to timezone-naive strings
                def convert_datetime_obj(val):
                    if isinstance(val, datetime):
                        # Remove timezone info if present and convert to string
                        if val.tzinfo is not None:
                            val = val.replace(tzinfo=None)
                        return val.strftime('%Y-%m-%d %H:%M:%S')
                    return val
                
                df_copy[col] = series.apply(convert_datetime_obj)
                print(f"   🕐 Converted datetime objects to strings in column: {col}")
                continue
            
            # Check if it's a string column that might be datetime-like
            sample = series.dropna().astype(str).head(10)
            if sample.empty:
                continue

            # Try parsing a sample to see if it's plausibly datetime
            parsed_sample = pd.to_datetime(sample, errors="coerce", utc=False)
            success_ratio = parsed_sample.notna().mean()

            if success_ratio >= 0.6:  # majority of sample looks like a date
                print(f"   📅 Found datetime-like object column: {col}")
                try:
                    parsed_full = pd.to_datetime(series, errors="coerce", utc=False)
                    # If tz-aware, strip tz to be Excel-friendly
                    if (
                        hasattr(parsed_full.dtype, "tz")
                        and parsed_full.dtype.tz is not None
                    ):
                        parsed_full = parsed_full.dt.tz_localize(None)
                        print(f"   🕐 Removed timezone from column: {col}")
                    df_copy[col] = parsed_full
                    print(f"   ✅ Converted {col} to datetime64[ns]")
                except Exception as e:
                    print(f"   ⚠️  Could not convert {col} to datetime: {e}")
                    # Leave as-is

    print("   ✅ Completed Excel datetime compatibility conversion")
    return df_copy


def apply_professional_styling(workbook, worksheet, sheet_name):
    """Apply a clean style to non-PL sheets.

    PL_* sheets are section-styled already; global styling can clash there.
    """
    try:
        if sheet_name.startswith("PL_"):
            return  # Skip styling for PL sheets

        header_fill = PatternFill(
            start_color="2F5597", end_color="2F5597", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        light_fill = PatternFill(
            start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        if worksheet.max_row > 0:
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border

        for row_num in range(2, worksheet.max_row + 1):
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = thin_border
                if row_num % 2 == 0:
                    cell.fill = light_fill

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(
                        max_length,
                        len(str(cell.value)) if cell.value is not None else 0,
                    )
                except Exception:
                    pass
            worksheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 10), 50
            )
        worksheet.freeze_panes = "A2"
    except Exception:
        pass


def format_date_for_excel(date_value):
    """Format date for Excel compatibility using DD-MMM-YY format."""
    if pd.isna(date_value) or date_value is None:
        return None

    try:
        if isinstance(date_value, str):
            # Parse string dates
            if date_value.strip() == "":
                return None
            parsed_date = pd.to_datetime(date_value, errors="coerce")
            if pd.isna(parsed_date):
                return date_value  # Return original if parsing fails
            date_value = parsed_date

        # Convert to datetime if it's a Timestamp
        if isinstance(date_value, pd.Timestamp):
            date_value = date_value.to_pydatetime()

        # Format as DD-MMM-YY (e.g., "15-Jan-25")
        if hasattr(date_value, "strftime"):
            return date_value.strftime("%d-%b-%y")

        return date_value
    except Exception:
        return date_value  # Return original value if any error occurs


def convert_all_dates_to_excel_format(df: pd.DataFrame) -> pd.DataFrame:
    """Convert ALL date columns in DataFrame to DD-MMM-YY format for Excel recognition.

    This function identifies date columns and converts them to DD-MMM-YY format
    that Excel recognizes as dates while maintaining string format for consistency.

    Args:
        df: DataFrame to process

    Returns:
        DataFrame with all dates formatted as DD-MMM-YY strings
    """
    if df.empty:
        return df

    df = df.copy()

    # Known date column patterns (case-insensitive)
    date_patterns = [
        "date",
        "due_date",
        "eta_date",
        "order_date",
        "delivery_date",
        "earliest_order_date",
        "created_date",
        "modified_date",
    ]

    # Find date columns by name patterns
    date_columns = []
    for col in df.columns:
        if any(pattern in col.lower() for pattern in date_patterns):
            date_columns.append(col)

    # Also check for datetime dtypes
    for col in df.columns:
        if col not in date_columns:
            if df[col].dtype.name.startswith(("datetime", "timestamp")):
                date_columns.append(col)
            elif df[col].dtype == "object":
                # Check if object column contains datetime-like values
                sample = df[col].dropna().head(5)
                if not sample.empty:
                    try:
                        parsed = pd.to_datetime(sample, errors="coerce")
                        if not parsed.isna().all():
                            date_columns.append(col)
                    except (ValueError, TypeError):
                        pass

    # Apply DD-MMM-YY formatting to all identified date columns
    for col in date_columns:
        print(f"   📅 Converting {col} to DD-MMM-YY format")
        df[col] = df[col].apply(format_date_for_excel)

    return df


def add_eta_date_column(df: pd.DataFrame) -> pd.DataFrame:
    """Add ETA_Date column calculated as Date + 10 working days.

    Args:
        df: DataFrame with Date column

    Returns:
        DataFrame with ETA_Date column added
    """
    if df.empty or "Date" not in df.columns:
        return df

    df = df.copy()

    # Convert Date column to datetime if it's not already
    if not pd.api.types.is_datetime64_any_dtype(df["Date"]):
        # Handle categorical data by converting to string first
        if pd.api.types.is_categorical_dtype(df["Date"]):
            df["Date"] = df["Date"].astype(str)
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    if df["Date"].isna().all():
        return df
    invalid_dates = df["Date"].isna()
    if invalid_dates.any():
        print(f"   ⚠️ {invalid_dates.sum()} invalid Date values; ETA_Date set to NaT")

    # Calculate ETA_Date as Date + 10 business days, skipping invalid entries
    df["ETA_Date"] = df["Date"].apply(
        lambda d: d + pd.offsets.BDay(10) if pd.notna(d) else pd.NaT
    )

    print("   📅 Added ETA_Date column (Date + 10 business days)")

    return df


def apply_date_number_formats(workbook):
    """Ensure known date columns are formatted as dates in Excel."""
    date_like_cols = {
        "Date",
        "Due_Date",
        "ETA_Date",
        "Earliest_Order_Date",
        "Order Date",
    }
    try:
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            if ws.max_row < 2:
                continue
            header = [cell.value for cell in ws[1]]
            for col_idx, name in enumerate(header, start=1):
                if name in date_like_cols:
                    for cell in ws.iter_cols(
                        min_col=col_idx,
                        max_col=col_idx,
                        min_row=2,
                        max_row=ws.max_row,
                    )[0]:
                        cell.number_format = "yyyy-mm-dd hh:mm:ss"
    except (KeyError, AttributeError, TypeError):
        pass


def load_conversion_factors() -> dict:
    """Load conversion factors from the product categorization Excel file."""
    try:
        base = Path(__file__).resolve().parent.parent.parent  # Go up to project root
        conversion_factors_path = base / "docs" / "product_categorization.xlsx"
        if conversion_factors_path.exists():
            df = pd.read_excel(
                conversion_factors_path, sheet_name="category_conversion_factors"
            )
            # Create a mapping from Product Category to conversion_factor
            conversion_map = dict(zip(df["Product Category"], df["conversion_factor"]))
            return conversion_map
        else:
            print(
                f"Warning: Conversion factors file not found at {conversion_factors_path}"
            )
            return {}
    except Exception as e:
        print(f"Error loading conversion factors: {e}")
        return {}


def load_product_category_mapping() -> dict:
    """Load mapping from Internal Reference (product code) to detailed Product Category."""
    try:
        base = Path(__file__).resolve().parent.parent.parent  # Go up to project root
        categorization_path = base / "docs" / "product_categorization.xlsx"
        if categorization_path.exists():
            df = pd.read_excel(categorization_path, sheet_name="product_info")
            # Create mapping from Internal Reference to Product Category
            # Use dropna() to avoid NaN keys and filter out empty strings
            df_clean = df.dropna(subset=["Internal Reference", "Product Category"])
            df_clean = df_clean[
                (df_clean["Internal Reference"].astype(str).str.strip() != "")
                & (df_clean["Product Category"].astype(str).str.strip() != "")
            ]

            product_mapping = dict(
                zip(
                    df_clean["Internal Reference"].astype(str).str.strip(),
                    df_clean["Product Category"].astype(str).str.strip(),
                )
            )
            print(f"Loaded {len(product_mapping)} product category mappings")
            return product_mapping
        else:
            print(
                f"Warning: Product categorization file not found at {categorization_path}"
            )
            return {}
    except Exception as e:
        print(f"Error loading product category mapping: {e}")
        return {}


def add_to_produce_column(df: pd.DataFrame, conversion_factors: dict) -> pd.DataFrame:
    """Add 'to_produce' column using proper product categorization mapping."""
    if df.empty:
        return df

    df = df.copy()

    if "Open_Qty" in df.columns:
        # Load product category mapping from Internal Reference to detailed categories
        product_category_mapping = load_product_category_mapping()

        # Try to map using Internal_Reference first (most accurate)
        if "Internal_Reference" in df.columns and product_category_mapping:
            print(
                f"   🔍 Mapping categories using Internal_Reference for {len(df)} items"
            )
            df["proper_category"] = (
                df["Internal_Reference"]
                .astype(str)
                .str.strip()
                .map(product_category_mapping)
            )

            # Count successful mappings
            successful_mappings = df["proper_category"].notna().sum()
            print(
                f"   ✅ Successfully mapped {successful_mappings}/{len(df)} products to categories"
            )

            # Show some examples of mappings for debugging
            if successful_mappings > 0:
                sample_mappings = df[df["proper_category"].notna()][
                    ["Internal_Reference", "proper_category"]
                ].head(3)
                print("   📋 Sample mappings:")
                for _, row in sample_mappings.iterrows():
                    print(
                        f"      {row['Internal_Reference']} → {row['proper_category']}"
                    )
        else:
            print("   ⚠️  No Internal_Reference column or product mapping available")
            df["proper_category"] = "Unknown"

        # Fallback to Odoo Category if no proper mapping found
        if "Category" in df.columns:
            df["proper_category"] = df["proper_category"].fillna(df["Category"])

        # Apply conversion factors using the proper categories (Context7 pattern: explicit numeric conversion)
        df["conversion_factor"] = (
            df["proper_category"].map(conversion_factors).fillna(1.0)
        )

        # Ensure Open_Qty is numeric before multiplication (fix for to_produce calculation)
        df["Open_Qty"] = pd.to_numeric(df["Open_Qty"], errors="coerce").fillna(0)
        df["to_produce"] = df["Open_Qty"] * df["conversion_factor"]

        # Report conversion factor statistics
        factor_stats = df["conversion_factor"].value_counts()
        print(f"   📊 Conversion factors applied: {dict(factor_stats)}")

        # Warn about products using default factor (1.0)
        default_factor_count = (df["conversion_factor"] == 1.0).sum()
        if default_factor_count > 0:
            print(
                f"   ⚠️  {default_factor_count} products using default conversion factor (1.0)"
            )

    else:
        # If Open_Qty column doesn't exist, set to_produce to 0
        df["to_produce"] = 0.0
        print("   ⚠️  No Open_Qty column found, setting to_produce to 0")

    return df


def create_pl_sheet_with_multiple_tables(writer, df: pd.DataFrame, sheet_name: str):
    """Create a PL sheet with multiple tables grouped by Item_Description, with 2 blank rows between tables."""
    if df.empty:
        df_export = ensure_excel_datetime_compatibility(df)
        df_export = convert_all_dates_to_excel_format(df_export)
        df_export.to_excel(writer, sheet_name=sheet_name, index=False)
        return

    # Group data by Item_Description
    if "Item_Description" not in df.columns:
        df_export = ensure_excel_datetime_compatibility(df)
        df_export = convert_all_dates_to_excel_format(df_export)
        df_export.to_excel(writer, sheet_name=sheet_name, index=False)
        return

    # Ensure Excel datetime compatibility and apply DD-MMM-YY formatting
    df = ensure_excel_datetime_compatibility(df)
    df = convert_all_dates_to_excel_format(df)

    # Handle missing/null Item_Description values
    df = df.copy()
    if pd.api.types.is_categorical_dtype(df["Item_Description"]):
        if "No Description" not in df["Item_Description"].cat.categories:
            df["Item_Description"] = df["Item_Description"].cat.add_categories(
                ["No Description"]
            )
        df["Item_Description"] = df["Item_Description"].fillna("No Description")
    else:
        df["Item_Description"] = df["Item_Description"].fillna("No Description")

    grouped = df.groupby("Item_Description", sort=True, observed=True)

    workbook = writer.book
    worksheet = workbook.create_sheet(sheet_name)

    current_row = 1
    table_counter = 1

    for idx, (item_description, group_df) in enumerate(grouped, start=1):
        # Section header
        worksheet.cell(
            row=current_row, column=1, value=f"Item Description: {item_description}"
        )
        header_cell = worksheet.cell(row=current_row, column=1)
        header_cell.font = Font(name="Calibri", size=12, bold=True, color="2F5597")
        current_row += 1

        # Column headers
        for col_idx, column_name in enumerate(group_df.columns, 1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=column_name)
            cell.fill = PatternFill(
                start_color="2F5597", end_color="2F5597", fill_type="solid"
            )
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin", color="D1D5DB"),
                right=Side(style="thin", color="D1D5DB"),
                top=Side(style="thin", color="D1D5DB"),
                bottom=Side(style="thin", color="D1D5DB"),
            )

        header_row = current_row
        current_row += 1

        # Data rows
        light_fill = PatternFill(
            start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        for row_idx, (_, row_data) in enumerate(group_df.iterrows()):
            for col_idx, value in enumerate(row_data, 1):
                # Convert pandas.Timestamp to python datetime for Excel
                if isinstance(value, pd.Timestamp):
                    value = value.to_pydatetime()
                cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = light_fill
            current_row += 1

        # Create Excel table for this group (robust beyond 26 columns)
        if len(group_df) > 0:
            end_col_letter = get_column_letter(len(group_df.columns))
            table_range = f"A{header_row}:{end_col_letter}{current_row - 1}"
            table_name = f"Table{table_counter}_{sheet_name.replace('PL_', '').replace(' ', '_')}"
            table_name = table_name.replace(" ", "_").replace("-", "_")[:50]

            try:
                table = Table(displayName=table_name, ref=table_range)
                style = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=True,
                )
                table.tableStyleInfo = style
                worksheet.add_table(table)
                table_counter += 1
            except Exception as e:
                print(f"Warning: Could not create table for {item_description}: {e}")

        # 2 blank rows between tables (except the last)
        if idx < grouped.ngroups:
            current_row += 2

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)) if cell.value else 0)
            except Exception:
                pass
        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 10), 50
        )

    worksheet.freeze_panes = "A2"


def prepare_pl_export(pl_data: pd.DataFrame) -> pd.DataFrame:
    """Prepare production line data for Excel export.

    Ensures a "Date" column exists (renaming "Order Date" if necessary),
    adds ETA_Date column, and selects the relevant columns for PL sheets.
    """
    if "Date" not in pl_data.columns and "Order Date" in pl_data.columns:
        pl_data = pl_data.rename(columns={"Order Date": "Date"})

    # Add ETA_Date column
    pl_data = add_eta_date_column(pl_data)

    pl_columns = [
        "Date",
        "Number",
        "Customer",
        "ETA_Date",
        "Internal_Reference",
        "Item_Description",
        "to_produce",
    ]

    available_pl_columns = [c for c in pl_columns if c in pl_data.columns]
    pl_export = pl_data[available_pl_columns].copy()

    print(f"   📋 PL sheet columns: {list(pl_export.columns)}")
    return pl_export


def filter_zero_to_produce(df: pd.DataFrame, column_name: str = "Total_to_Produce") -> pd.DataFrame:
    """Filter out rows where the specified to_produce column equals 0.
    
    Args:
        df: DataFrame to filter
        column_name: Name of the column to check for zero values (default: "Total_to_Produce")
        
    Returns:
        Filtered DataFrame with zero-value rows removed
    """
    if df.empty or column_name not in df.columns:
        return df
    
    original_count = len(df)
    # Use pandas filtering pattern to remove rows where column equals 0
    filtered_df = df[df[column_name] != 0].copy()
    filtered_count = len(filtered_df)
    
    if original_count != filtered_count:
        print(f"   🧹 Filtered out {original_count - filtered_count} rows with {column_name} == 0 (kept {filtered_count} rows)")
    
    return filtered_df


def create_to_produce_sheets(df: pd.DataFrame) -> tuple:
    """Create separate to_produce sheets for Food Sales and Detergent Sales categories."""
    food_sales_df = pd.DataFrame()
    detergent_sales_df = pd.DataFrame()

    # Check for both Sample_Category (current) and Category (fallback)
    category_column = None
    if "Sample_Category" in df.columns:
        category_column = "Sample_Category"
    elif "Category" in df.columns:
        category_column = "Category"

    if category_column:
        print(f"Using {category_column} column for category filtering")

        # Food Sales
        food_mask = (
            df[category_column]
            .astype(str)
            .str.contains("Food Sales", case=False, na=False)
        )
        if food_mask.any():
            food_sales_df = df[food_mask].copy()
            print(
                f"Created Food Sales to_produce sheet with {len(food_sales_df)} products"
            )

        # Detergent Sales
        detergent_mask = (
            df[category_column]
            .astype(str)
            .str.contains("Detergent Sales", case=False, na=False)
        )
        if detergent_mask.any():
            detergent_sales_df = df[detergent_mask].copy()
            print(
                f"Created Detergent Sales to_produce sheet with {len(detergent_sales_df)} products"
            )
    else:
        print(
            "Neither Sample_Category nor Category column found in data - cannot create category-specific to_produce sheets"
        )

    return food_sales_df, detergent_sales_df


def generate_quickbooks_excel_report(
    file_path: str,
    output_dir: str = "output",
) -> Tuple[Optional[str], Optional[str]]:
    """Generate comprehensive QuickBooks Excel report with detailed analysis.
    
    Args:
        file_path: Path to QuickBooks CSV file
        output_dir: Directory where Excel and JSON reports will be saved (default: "output")
        
    Returns:
        Tuple of (excel_file_path, json_file_path) or (None, None) if generation fails
    """
    # Extract orders from CSV
    orders_df = extract_sales_orders_from_csv(file_path)

    # Get line items detail to calculate proper totals
    print(f"\n📊 Fetching line items detail for {len(orders_df)} orders...")
    line_items_detail = get_odoo_sales_order_line_items(orders_df, logger=_NoopLogger())
    print(f"   ✅ Retrieved {len(line_items_detail)} line items")
    if not line_items_detail.empty:
        print(f"   📋 Available columns: {list(line_items_detail.columns)}")
        if "Production_Line" in line_items_detail.columns:
            production_lines = line_items_detail["Production_Line"].value_counts()
            print(f"   🏭 Production lines found: {dict(production_lines)}")
        else:
            print("   ⚠️  No Production_Line column in line_items_detail")
    else:
        print("   ⚠️  No line items detail retrieved!")

    # Restrict line items strictly to orders present in the CSV (by Number)
    try:
        if (
            not orders_df.empty
            and "order_number" in orders_df.columns
            and not line_items_detail.empty
            and "Number" in line_items_detail.columns
        ):
            allowed_numbers = set(
                orders_df["order_number"].dropna().map(lambda x: str(x).strip())
            )
            print(f"   🔍 CSV order numbers: {allowed_numbers}")

            odoo_numbers = set(
                line_items_detail["Number"]
                .dropna()
                .map(lambda x: str(x).strip())
                .unique()
            )
            print(f"   🔍 Odoo order numbers (sample): {list(odoo_numbers)[:10]}")

            line_items_detail = line_items_detail.copy()
            line_items_detail["Number"] = line_items_detail["Number"].map(
                lambda x: str(x).strip()
            )

            overlap = allowed_numbers.intersection(odoo_numbers)
            print(f"   🔍 Matching order numbers: {overlap}")

            if not overlap:
                print(
                    "   ⚠️  No matching order numbers found - keeping all line items for debugging"
                )
            else:
                line_items_detail = line_items_detail[
                    line_items_detail["Number"].isin(allowed_numbers)
                ].copy()
                print(
                    f"   ✅ Filtered to {len(line_items_detail)} line items matching CSV orders"
                )
    except Exception as e:
        print(f"   ⚠️  Error in order filtering: {e}")
        pass

    # Create orders summary from aggregated line items data (with Odoo dates)
    # Using Context7 pandas groupby patterns for proper aggregation
    if not line_items_detail.empty and "Number" in line_items_detail.columns:
        print(
            f"   📊 Creating Orders_Summary from {len(line_items_detail)} line items..."
        )

        # Ensure numeric columns are properly typed before aggregation
        numeric_cols = ["Qty", "Open_Qty", "Open_Balance"]
        for col in numeric_cols:
            if col in line_items_detail.columns:
                line_items_detail[col] = pd.to_numeric(
                    line_items_detail[col], errors="coerce"
                ).fillna(0)

        order_groups = line_items_detail.groupby("Number", dropna=True, observed=True)
        orders_summary = order_groups.agg(
            order_date=("Order Date", "min"),
            delivery_date=("Due_Date", "min"),
            customer_name=("Customer", "first"),
            memo=("Memo", "first"),
            total_qty=("Qty", "sum"),
            open_qty=("Open_Qty", "sum"),
            open_balance=("Open_Balance", "sum"),
        ).reset_index()

        orders_summary = orders_summary.rename(
            columns={
                "Number": "Number",
                "order_date": "Date",
                "delivery_date": "Due_Date",
                "customer_name": "Customer",
                "total_qty": "Total_Item_Qty",
                "open_qty": "Open_Qty",
                "open_balance": "Open_Balance",
            }
        )

        print(f"   ✅ Created Orders_Summary with {len(orders_summary)} orders")
        print(f"   📊 Total_Item_Qty sum: {orders_summary['Total_Item_Qty'].sum():.2f}")
        print(f"   📊 Open_Qty sum: {orders_summary['Open_Qty'].sum():.2f}")
    else:
        # Fallback to original CSV data if no line items
        orders_summary = orders_df.copy()

        # Ensure Date column is properly formatted first
        if "order_date" in orders_summary.columns:
            orders_summary["Date"] = pd.to_datetime(
                orders_summary["order_date"], errors="coerce"
            )
        elif "Order Date" in orders_summary.columns:
            orders_summary["Date"] = pd.to_datetime(
                orders_summary["Order Date"], errors="coerce"
            )

        column_mapping = {
            "order_number": "Number",
            "customer_name": "Customer",
            "delivery_date": "Due_Date",
            "memo": "Memo",
            "open_balance": "Open_Balance",
        }
        existing_mapping = {
            k: v for k, v in column_mapping.items() if k in orders_summary.columns
        }
        orders_summary = orders_summary.rename(columns=existing_mapping)

    # Skip redundant calculation - quantities already calculated above in the main aggregation
    # The previous logic was overwriting the correct aggregated values with defaults
    # Context7 pattern: avoid redundant operations that can introduce errors
    if "Total_Item_Qty" not in orders_summary.columns:
        print("   ⚠️  No Total_Item_Qty found, setting defaults")
        orders_summary["Total_Item_Qty"] = 1.0
    if "Open_Qty" not in orders_summary.columns:
        print("   ⚠️  No Open_Qty found, setting defaults")
        orders_summary["Open_Qty"] = 1.0

    # Ensure date columns populated from line items if available
    if not line_items_detail.empty and "Number" in line_items_detail.columns:
        if "Date" not in orders_summary.columns or orders_summary["Date"].isna().all():
            date_mapping = line_items_detail.groupby("Number")["Order Date"].first()
            orders_summary = (
                orders_summary.set_index("Number")
                .fillna({"Date": date_mapping})
                .reset_index()
            )

        if (
            "Due_Date" not in orders_summary.columns
            or orders_summary["Due_Date"].isna().all()
        ):
            due_date_mapping = line_items_detail.groupby("Number")["Due_Date"].first()
            orders_summary = (
                orders_summary.set_index("Number")
                .fillna({"Due_Date": due_date_mapping})
                .reset_index()
            )

    # Add Formula info (BOM)
    if not line_items_detail.empty and "Item_Code" in line_items_detail.columns:
        item_codes = [
            c
            for c in line_items_detail["Item_Code"].dropna().unique().tolist()
            if str(c).strip()
        ]
        if item_codes:
            formula_info = get_formula_info_from_bom(item_codes, logger=_NoopLogger())
            line_items_detail["Formula_Name"] = line_items_detail["Item_Code"].map(
                lambda x: (
                    formula_info.get(x, {}).get("formula_name", "INFO MISSING")
                    if x
                    else "INFO MISSING"
                )
            )
            line_items_detail["BOM_Reference"] = line_items_detail["Item_Code"].map(
                lambda x: (
                    formula_info.get(x, {}).get("reference", "INFO MISSING")
                    if x
                    else "INFO MISSING"
                )
            )
        else:
            line_items_detail["Formula_Name"] = "INFO MISSING"
            line_items_detail["BOM_Reference"] = "INFO MISSING"
    elif not line_items_detail.empty:
        line_items_detail["Formula_Name"] = "INFO MISSING"
        line_items_detail["BOM_Reference"] = "INFO MISSING"

    # Filter out rows with empty Internal_Reference (using Context7 pandas filtering patterns)
    if (
        not line_items_detail.empty
        and "Internal_Reference" in line_items_detail.columns
    ):
        print("   🔍 Filtering out rows with empty Internal_Reference...")
        original_count = len(line_items_detail)

        # Filter using Context7 pandas patterns: notna() and string filtering
        line_items_detail = line_items_detail[
            line_items_detail["Internal_Reference"].notna()
            & (line_items_detail["Internal_Reference"].astype(str).str.strip() != "")
            & (line_items_detail["Internal_Reference"].astype(str).str.strip() != "nan")
        ].copy()

        filtered_count = len(line_items_detail)
        dropped_count = original_count - filtered_count
        if dropped_count > 0:
            print(
                f"   ✅ Dropped {dropped_count} rows with empty Internal_Reference ({original_count} → {filtered_count})"
            )
        else:
            print("   ✅ No rows with empty Internal_Reference found")

    # Fix Category column using proper categorization from product_categorization.xlsx
    # Context7 pattern: Use .map() method for efficient column replacement
    if (
        not line_items_detail.empty
        and "Internal_Reference" in line_items_detail.columns
    ):
        print("   🔄 Fixing Category column using proper categorization...")

        # Load the proper category mapping from the Excel file
        product_category_mapping = load_product_category_mapping()

        if product_category_mapping:
            # Map Internal_Reference to proper categories using Context7 .map() pattern
            original_category_count = (
                line_items_detail["Category"].nunique()
                if "Category" in line_items_detail.columns
                else 0
            )

            # Replace the Category column with proper categorization
            line_items_detail["Category"] = (
                line_items_detail["Internal_Reference"]
                .astype(str)
                .str.strip()
                .map(product_category_mapping)
            )

            # Fill any unmapped categories with "Unknown" (Context7 pattern: handle missing values)
            line_items_detail["Category"] = line_items_detail["Category"].fillna(
                "Unknown"
            )

            # Report the mapping results
            new_category_count = line_items_detail["Category"].nunique()
            mapped_count = (line_items_detail["Category"] != "Unknown").sum()

            print(
                f"   ✅ Category column updated: {mapped_count}/{len(line_items_detail)} items mapped"
            )
            print(
                f"   📊 Categories: {original_category_count} → {new_category_count} unique categories"
            )

            # Show sample of new categories
            category_samples = line_items_detail["Category"].value_counts().head(5)
            print(f"   📋 Top categories: {dict(category_samples)}")
        else:
            print(
                "   ⚠️  No product category mapping available - keeping original Category column"
            )

    # Load conversion factors and add to_produce
    conversion_factors = load_conversion_factors()
    line_items_detail = add_to_produce_column(line_items_detail, conversion_factors)

    orders_summary = ensure_excel_datetime_compatibility(
        optimize_dataframe_dtypes(orders_summary)
    )
    line_items_detail = ensure_excel_datetime_compatibility(
        optimize_dataframe_dtypes(line_items_detail)
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"{output_dir}/QuickBooks_Sales_Orders_Report_{timestamp}.xlsx"
    json_filename = f"{output_dir}/QuickBooks_Sales_Orders_Report_{timestamp}.json"
    os.makedirs(output_dir, exist_ok=True)

    with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
        # Orders_Summary sheet
        summary_columns = [
            "Date",
            "Number",
            "Customer",
            "Due_Date",
            "ETA_Date",
            "Memo",
            "Total_Item_Qty",
            "Open_Qty",
            "Open_Balance",
        ]
        # Add ETA_Date column to orders_summary
        orders_summary = add_eta_date_column(orders_summary)

        available_columns = [
            col for col in summary_columns if col in orders_summary.columns
        ]
        orders_summary_export = (
            orders_summary.copy()
            if not available_columns
            else orders_summary[available_columns].copy()
        )

        # Keep datetimes as datetimes; ensure numeric types
        numeric_columns = ["Total_Item_Qty", "Open_Qty", "Open_Balance"]
        for col in numeric_columns:
            if col in orders_summary_export.columns:
                orders_summary_export[col] = pd.to_numeric(
                    orders_summary_export[col], errors="coerce"
                ).fillna(0)
                if col in ["Total_Item_Qty", "Open_Qty"]:
                    orders_summary_export[col] = (
                        orders_summary_export[col].round(0).astype(int)
                    )
                else:
                    orders_summary_export[col] = orders_summary_export[col].round(2)

        orders_summary_export = convert_all_dates_to_excel_format(orders_summary_export)
        orders_summary_export.to_excel(writer, sheet_name="Orders_Summary", index=False)

        # Line_Items_Detail sheet
        line_items_detail_columns = [
            "Date",  # (renamed from "Order Date" if needed)
            "Number",
            "Customer",
            "Due_Date",
            "ETA_Date",
            "Internal_Reference",
            "Item_Description",
            "Category",
            "Production_Line",
            "Formula_Name",
            "Memo",
            "to_produce",
        ]
        line_items_export = line_items_detail.copy()
        if (
            "Order Date" in line_items_export.columns
            and "Date" not in line_items_export.columns
        ):
            line_items_export = line_items_export.rename(columns={"Order Date": "Date"})

        # Add ETA_Date column to line_items_export
        line_items_export = add_eta_date_column(line_items_export)

        available_line_items_columns = [
            c for c in line_items_detail_columns if c in line_items_export.columns
        ]
        line_items_export = line_items_export[available_line_items_columns].copy()
        print(f"   📋 Line_Items_Detail columns: {list(line_items_export.columns)}")

        line_items_export = ensure_excel_datetime_compatibility(line_items_export)
        line_items_export = convert_all_dates_to_excel_format(line_items_export)
        line_items_export.to_excel(writer, sheet_name="Line_Items_Detail", index=False)

        # Orders_by_Customer sheet (fixed Open_Qty math)
        customer_summary = (
            orders_summary.groupby("Customer", observed=True)
            .agg(
                Order_Count=("Number", "count"),
                Total_Item_Qty=(
                    ("Total_Item_Qty", "sum")
                    if "Total_Item_Qty" in orders_summary.columns
                    else ("Number", "size")
                ),
                Open_Qty=(
                    ("Open_Qty", "sum")
                    if "Open_Qty" in orders_summary.columns
                    else ("Total_Item_Qty", "sum")
                ),
                Earliest_Order_Date=("Date", "min"),
                Open_Balance=("Open_Balance", "sum"),
            )
            .reset_index()
            .sort_values("Open_Balance", ascending=False)
        )
        customer_summary = ensure_excel_datetime_compatibility(customer_summary)
        customer_summary = convert_all_dates_to_excel_format(customer_summary)
        customer_summary.to_excel(writer, sheet_name="Orders_by_Customer", index=False)

        # Category counts for summary
        food_sales_count = 0
        detergent_sales_count = 0
        if not line_items_detail.empty and "Category" in line_items_detail.columns:
            food_sales_count = (
                line_items_detail["Category"]
                .astype(str)
                .str.contains("Food Sales", case=False, na=False)
                .sum()
            )
            detergent_sales_count = (
                line_items_detail["Category"]
                .astype(str)
                .str.contains("Detergent Sales", case=False, na=False)
                .sum()
            )

        total_open_qty_val = (
            orders_summary["Open_Qty"].sum()
            if "Open_Qty" in orders_summary.columns
            else orders_summary["Total_Item_Qty"].sum()
        )

        summary_df = pd.DataFrame(
            {
                "Metric": [
                    "Total Open Orders",
                    "Total Customers with Open Orders",
                    "Total Open Balance",
                    "Total Open Quantity",
                    "Average Order Value",
                    "Largest Order Value",
                    "Food Sales Products",
                    "Detergent Sales Products",
                    "Report Generated",
                    "Source File",
                ],
                "Value": [
                    f"{len(orders_summary):,}",
                    f"{orders_summary['Customer'].nunique():,}",
                    f"${orders_summary['Open_Balance'].sum():,.2f}",
                    f"{total_open_qty_val:,.0f}",
                    f"${orders_summary['Open_Balance'].mean():,.2f}",
                    f"${orders_summary['Open_Balance'].max():,.2f}",
                    f"{food_sales_count:,}",
                    f"{detergent_sales_count:,}",
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    os.path.basename(file_path),
                ],
            }
        )
        summary_df = ensure_excel_datetime_compatibility(summary_df)
        summary_df = convert_all_dates_to_excel_format(summary_df)
        summary_df.to_excel(writer, sheet_name="Executive_Summary", index=False)

        # to_produce summary grouped by Formula_Name (+ split by category)
        if (
            not line_items_detail.empty
            and "Formula_Name" in line_items_detail.columns
            and "to_produce" in line_items_detail.columns
        ):
            valid_formulas = line_items_detail[
                (line_items_detail["Formula_Name"].notna())
                & (line_items_detail["Formula_Name"] != "INFO MISSING")
                & (line_items_detail["Formula_Name"] != "")
            ].copy()

            if not valid_formulas.empty:
                to_produce_summary = (
                    valid_formulas.groupby("Formula_Name", observed=True)
                    .agg(
                        Total_to_Produce=("to_produce", "sum"),
                        Total_Open_Qty=("Open_Qty", "sum"),
                        Sample_Category=("Category", "first"),
                        Sample_Production_Line=("Production_Line", "first"),
                        Sample_BOM_Reference=("BOM_Reference", "first"),
                    )
                    .reset_index()
                    .sort_values("Total_to_Produce", ascending=False)
                )

                to_produce_summary["Total_to_Produce"] = to_produce_summary[
                    "Total_to_Produce"
                ].round(2)
                to_produce_summary["Total_Open_Qty"] = to_produce_summary[
                    "Total_Open_Qty"
                ].round(2)

                food_sales_df, detergent_sales_df = create_to_produce_sheets(
                    to_produce_summary
                )

                # Filter out rows with Total_to_Produce == 0 from Food Sales sheet
                food_sales_df = filter_zero_to_produce(food_sales_df, "Total_to_Produce")

                if not food_sales_df.empty:
                    # Drop columns using Context7 pandas patterns - drop with errors='ignore' for safety
                    food_sales_export = food_sales_df.drop(
                        columns=[
                            "Total_Open_Qty",
                            "Sample_Category",
                            "Sample_BOM_Reference",
                        ],
                        errors="ignore",
                    )
                    food_sales_export = ensure_excel_datetime_compatibility(
                        food_sales_export
                    )
                    food_sales_export = convert_all_dates_to_excel_format(
                        food_sales_export
                    )
                    food_sales_export.to_excel(
                        writer, sheet_name="to_produce_Food_Sales", index=False
                    )
                    print(
                        "Dropped Total_Open_Qty, Sample_Category and Sample_BOM_Reference from to_produce_Food_Sales sheet"
                    )

                # Filter out rows with Total_to_Produce == 0 from Detergent Sales sheet
                detergent_sales_df = filter_zero_to_produce(detergent_sales_df, "Total_to_Produce")

                if not detergent_sales_df.empty:
                    # Drop columns using Context7 pandas patterns - drop with errors='ignore' for safety
                    detergent_sales_export = detergent_sales_df.drop(
                        columns=[
                            "Total_Open_Qty",
                            "Sample_Category",
                            "Sample_BOM_Reference",
                        ],
                        errors="ignore",
                    )
                    detergent_sales_export = ensure_excel_datetime_compatibility(
                        detergent_sales_export
                    )
                    detergent_sales_export = convert_all_dates_to_excel_format(
                        detergent_sales_export
                    )
                    detergent_sales_export.to_excel(
                        writer, sheet_name="to_produce_Detergent_Sales", index=False
                    )
                    print(
                        "Dropped Total_Open_Qty, Sample_Category and Sample_BOM_Reference from to_produce_Detergent_Sales sheet"
                    )

                # Filter out rows with Total_to_Produce == 0 from main to_produce sheet
                to_produce_summary_filtered = filter_zero_to_produce(to_produce_summary, "Total_to_Produce")
                
                # Drop specified columns from main to_produce sheet using Context7 pandas patterns
                to_produce_export = to_produce_summary_filtered.drop(
                    columns=["Total_Open_Qty", "Sample_BOM_Reference"], errors="ignore"
                )
                to_produce_export = ensure_excel_datetime_compatibility(
                    to_produce_export
                )
                to_produce_export = convert_all_dates_to_excel_format(to_produce_export)
                to_produce_export.to_excel(writer, sheet_name="to produce", index=False)
                print(
                    "Dropped Total_Open_Qty and Sample_BOM_Reference from to_produce sheet"
                )

        # Line_Items_Ready sheet removed - not needed right now
        # This sheet was showing items with open_balance == 0
        print("   ℹ️  Skipping Line_Items_Ready sheet - not needed right now")

        # Production Line sheets (multi-table by Item_Description)
        print("\n📊 Creating Production Line sheets...")
        if "Production_Line" in line_items_detail.columns:
            production_lines = line_items_detail["Production_Line"].dropna().unique()
            print(
                f"   Found {len(production_lines)} production lines: {list(production_lines)}"
            )

            for production_line in production_lines:
                if pd.isna(production_line) or production_line == "":
                    continue

                pl_data = line_items_detail[
                    line_items_detail["Production_Line"] == production_line
                ].copy()
                if pl_data.empty:
                    print(f"   ⚠️  No data for production line: {production_line}")
                    continue

                safe_sheet_name = (
                    str(production_line).replace("/", "_").replace("\\", "_")[:31]
                )
                print(
                    f"   📋 Creating PL_{safe_sheet_name} sheet with {len(pl_data)} items"
                )

                pl_export = prepare_pl_export(pl_data)
                # Filter out rows with to_produce == 0 from PL sheets
                pl_export = filter_zero_to_produce(pl_export, "to_produce")
                
                if not pl_export.empty:
                    pl_export = ensure_excel_datetime_compatibility(pl_export)
                    pl_export = convert_all_dates_to_excel_format(pl_export)
                    create_pl_sheet_with_multiple_tables(
                        writer, pl_export, f"PL_{safe_sheet_name}"
                    )
                    print(f"   ✅ Created PL_{safe_sheet_name} sheet")
                else:
                    print(f"   ⚠️  No data after preparation for PL_{safe_sheet_name}")
        else:
            print("   ⚠️  No Production_Line column found in line_items_detail")

    # Post-process workbook: move sheets, style, and apply date formats
    try:
        workbook = load_workbook(excel_filename)

        # Move Executive_Summary sheet to first position
        if "Executive_Summary" in workbook.sheetnames:
            exec_summary_sheet = workbook["Executive_Summary"]
            workbook.move_sheet(
                exec_summary_sheet, offset=-workbook.index(exec_summary_sheet)
            )
            print("Moved Executive_Summary sheet to first position")

        for sheet_name in workbook.sheetnames:
            apply_professional_styling(workbook, workbook[sheet_name], sheet_name)

        apply_date_number_formats(workbook)

        workbook.save(excel_filename)
        workbook.close()
    except Exception as e:
        print(f"Error during post-processing: {e}")
        pass

    export_data = {
        "orders_summary": orders_summary.to_dict("records"),
        "line_items_detail": line_items_detail.to_dict("records"),
        "metadata": {
            "export_timestamp": datetime.now().isoformat(),
            "total_open_orders": len(orders_summary),
            "total_open_line_items": len(line_items_detail),
            "data_source": "QuickBooks CSV Export",
            "export_format_version": "1.0",
            "generator": "quickbooks.excel_export",
            "report_type": "Open Sales Orders",
            "source_file": file_path,
        },
    }
    with open(json_filename, "w", encoding="utf-8") as json_file:
        json.dump(export_data, json_file, indent=2, ensure_ascii=False, default=str)

    return excel_filename, json_filename


class _NoopLogger:
    def info(self, *args, **kwargs):
        pass

    def warning(self, *args, **kwargs):
        pass

    def error(self, *args, **kwargs):
        pass
