"""Diagnostics service for QuickBooks Auto Reporter.

Provides comprehensive QuickBooks connectivity diagnostics and
troubleshooting recommendations.
"""

import json
import os
import sys
import datetime as dt
from typing import Dict, Any, Tuple

from ..adapters.quickbooks.connection import initialize_com, open_connection, try_begin_session, cleanup_com
from ..adapters.quickbooks.error_handler import get_user_friendly_error
from ..config import DEFAULT_OUT_DIR
from ..utils.logging_utils import log_diagnostic, log_info, log_success, log_error, log_separator


def check_quickbooks_installation() -> Tuple[bool, list]:
    """Check if QuickBooks Desktop is installed on the system.
    
    Returns:
        Tuple of (is_installed, details_list)
    """
    try:
        import winreg  # type: ignore
        
        # Check for QuickBooks in registry
        qb_paths = []
        registry_keys = [
            r"SOFTWARE\Intuit\QuickBooks",
            r"SOFTWARE\WOW6432Node\Intuit\QuickBooks",
        ]
        
        for key_path in registry_keys:
            try:
                with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, key_path) as key:
                    i = 0
                    while True:
                        try:
                            subkey_name = winreg.EnumKey(key, i)
                            if subkey_name.startswith(('QB', 'QuickBooks')):
                                qb_paths.append(f"{key_path}\\{subkey_name}")
                            i += 1
                        except WindowsError:
                            break
            except FileNotFoundError:
                continue
        
        return len(qb_paths) > 0, qb_paths
    except Exception as e:
        return False, [f"Registry check failed: {e}"]


def check_sdk_installation() -> Tuple[bool, str]:
    """Check if QuickBooks SDK is properly installed and registered.
    
    Returns:
        Tuple of (is_installed, details_message)
    """
    try:
        import pythoncom  # type: ignore
        
        # Try to create the COM object without initializing
        pythoncom.CoInitialize()
        try:
            # Check if QBXMLRP2.RequestProcessor is registered
            clsid = pythoncom.CLSIDFromProgID("QBXMLRP2.RequestProcessor")
            return True, f"SDK found with CLSID: {clsid}"
        except pythoncom.com_error as e:
            return False, f"COM registration error: {e}"
        finally:
            pythoncom.CoUninitialize()
    except Exception as e:
        return False, f"SDK check failed: {e}"


def test_com_object_creation() -> Tuple[bool, str]:
    """Test COM object creation.
    
    Returns:
        Tuple of (success, message)
    """
    try:
        rp = initialize_com()
        cleanup_com(rp)
        return True, "COM object creation successful"
    except Exception as e:
        return False, f"COM object creation failed: {e}"


def test_quickbooks_connection() -> Tuple[bool, str]:
    """Test connection to QuickBooks.
    
    Returns:
        Tuple of (success, message)
    """
    rp = None
    try:
        rp = initialize_com()
        open_connection(rp)
        ticket, info = try_begin_session(rp)
        
        # Clean up session
        try:
            rp.EndSession(ticket)
        except Exception:
            pass
        
        return True, f"Connection successful. Company: {info.get('CompanyFileName', 'Unknown')}"
    except Exception as e:
        error_info = get_user_friendly_error(e)
        return False, f"Connection failed: {error_info['message']}"
    finally:
        cleanup_com(rp)


def diagnose_quickbooks_connection(out_dir: str = None) -> Dict[str, Any]:
    """Run comprehensive QuickBooks connectivity diagnostics.
    
    Args:
        out_dir: Output directory for diagnostics (uses default if None)
        
    Returns:
        Dictionary containing diagnostic results
    """
    if out_dir is None:
        out_dir = DEFAULT_OUT_DIR
    
    os.makedirs(out_dir, exist_ok=True)
    
    diagnostics = {
        "timestamp": dt.datetime.now().isoformat(),
        "system_info": {},
        "quickbooks_installation": {},
        "sdk_installation": {},
        "connectivity_test": {},
        "recommendations": []
    }
    
    log_diagnostic("Running QuickBooks diagnostics...", out_dir)
    
    # System information
    try:
        diagnostics["system_info"] = {
            "platform": sys.platform,
            "python_version": sys.version,
            "architecture": "64-bit" if sys.maxsize > 2**32 else "32-bit"
        }
    except Exception as e:
        diagnostics["system_info"]["error"] = str(e)
    
    # Check QuickBooks installation
    qb_installed, qb_info = check_quickbooks_installation()
    diagnostics["quickbooks_installation"] = {
        "installed": qb_installed,
        "details": qb_info,
        "status": "✅ Found" if qb_installed else "❌ Not Found"
    }
    
    # Check SDK installation
    sdk_installed, sdk_info = check_sdk_installation()
    diagnostics["sdk_installation"] = {
        "installed": sdk_installed,
        "details": sdk_info,
        "status": "✅ Registered" if sdk_installed else "❌ Not Registered"
    }
    
    # Test connectivity
    log_diagnostic("Testing COM object creation...", out_dir)
    com_success, com_msg = test_com_object_creation()
    diagnostics["connectivity_test"]["com_object_creation"] = {
        "success": com_success,
        "message": com_msg,
        "status": "✅ Success" if com_success else "❌ Failed"
    }
    
    if com_success:
        log_diagnostic("Testing QuickBooks connection...", out_dir)
        conn_success, conn_msg = test_quickbooks_connection()
        diagnostics["connectivity_test"]["connection_test"] = {
            "success": conn_success,
            "message": conn_msg,
            "status": "✅ Success" if conn_success else "❌ Failed"
        }
        
        if not conn_success:
            # Add error details
            try:
                rp = initialize_com()
                try:
                    open_connection(rp)
                    try_begin_session(rp)
                except Exception as conn_e:
                    error_info = get_user_friendly_error(conn_e)
                    diagnostics["connectivity_test"]["connection_error"] = error_info
            finally:
                cleanup_com(rp)
    
    # Generate recommendations
    if not qb_installed:
        diagnostics["recommendations"].append("Install QuickBooks Desktop from Intuit")
    
    if not sdk_installed:
        diagnostics["recommendations"].append("Download and install QuickBooks SDK from Intuit Developer website")
        diagnostics["recommendations"].append("Run the application as Administrator after SDK installation")
    
    if qb_installed and sdk_installed and com_success:
        if not diagnostics["connectivity_test"].get("connection_test", {}).get("success", False):
            error_type = diagnostics["connectivity_test"].get("connection_error", {}).get("error_type")
            if error_type == "ACCESS_DENIED":
                diagnostics["recommendations"].append("Try running as Administrator")
            elif error_type == "FILE_NOT_FOUND":
                diagnostics["recommendations"].append("Verify QuickBooks company file path")
            elif error_type == "CONNECTION_ERROR":
                diagnostics["recommendations"].append("Make sure QuickBooks Desktop is running in single-user mode")
        else:
            diagnostics["recommendations"].append("QuickBooks and SDK appear to be working correctly")
    
    # Save diagnostics to JSON
    diag_file = os.path.join(out_dir, "quickbooks_diagnostics.json")
    with open(diag_file, "w", encoding="utf-8") as f:
        json.dump(diagnostics, f, indent=2)
    
    # Create Excel diagnostic report
    create_diagnostic_excel_report(diagnostics, out_dir)
    
    log_success(f"Diagnostics completed. Report saved to: {diag_file}", out_dir)
    return diagnostics


def create_diagnostic_excel_report(diagnostics: Dict[str, Any], out_dir: str) -> None:
    """Create an Excel diagnostic report.
    
    Args:
        diagnostics: Diagnostic results dictionary
        out_dir: Output directory for the report
    """
    try:
        excel_path = os.path.join(out_dir, "QuickBooks_Diagnostic_Report.xlsx")
        
        # Prepare diagnostic data for Excel
        summary_data = [
            ["QuickBooks Diagnostic Report", ""],
            ["Generated", diagnostics.get("timestamp", "")],
            ["", ""],
            ["Component", "Status"],
            ["QuickBooks Desktop", diagnostics["quickbooks_installation"]["status"]],
            ["QuickBooks SDK", diagnostics["sdk_installation"]["status"]],
            ["", ""],
            ["System Information", ""],
            ["Platform", diagnostics["system_info"].get("platform", "Unknown")],
            ["Python Version", diagnostics["system_info"].get("python_version", "Unknown")],
            ["Architecture", diagnostics["system_info"].get("architecture", "Unknown")],
        ]
        
        # Add connectivity test results
        if "connectivity_test" in diagnostics:
            summary_data.extend([
                ["", ""],
                ["Connectivity Tests", ""],
                ["COM Object Creation", diagnostics["connectivity_test"].get("com_object_creation", {}).get("status", "Not tested")],
            ])
            
            conn_test = diagnostics["connectivity_test"].get("connection_test")
            if conn_test:
                summary_data.append(["Connection Test", conn_test.get("status", "Not tested")])
        
        # Add recommendations
        if diagnostics.get("recommendations"):
            summary_data.extend([
                ["", ""],
                ["Recommendations", ""],
            ])
            for i, rec in enumerate(diagnostics["recommendations"], 1):
                summary_data.append([f"{i}.", rec])
        
        # Try to create Excel file with openpyxl
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Diagnostic Report"
            
            # Write data with basic formatting
            for row_idx, (col1, col2) in enumerate(summary_data, 1):
                ws.cell(row=row_idx, column=1, value=col1)
                ws.cell(row=row_idx, column=2, value=col2)
                
                # Format header
                if row_idx == 1:
                    cell = ws.cell(row=row_idx, column=1)
                    cell.font = Font(bold=True, size=14)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF", size=14)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_path)
            log_success(f"Excel diagnostic report created: {os.path.basename(excel_path)}", out_dir)
            
        except ImportError:
            log_error("openpyxl not available for Excel diagnostic report", out_dir)
        except Exception as excel_error:
            log_error(f"Excel diagnostic report creation failed: {excel_error}", out_dir)
    
    except Exception as e:
        log_error(f"Diagnostic Excel report creation failed: {e}", out_dir)


def print_diagnostics_summary(diagnostics: Dict[str, Any]) -> None:
    """Print a summary of diagnostics to console.
    
    Args:
        diagnostics: Diagnostic results dictionary
    """
    print("\nDIAGNOSTIC RESULTS:")
    print("=" * 30)
    print(f"QuickBooks Desktop: {diagnostics['quickbooks_installation']['status']}")
    print(f"QuickBooks SDK: {diagnostics['sdk_installation']['status']}")
    
    if "connectivity_test" in diagnostics:
        com_test = diagnostics["connectivity_test"].get("com_object_creation", {})
        print(f"COM Object Creation: {com_test.get('status', 'Not tested')}")
        
        conn_test = diagnostics["connectivity_test"].get("connection_test")
        if conn_test:
            print(f"Connection Test: {conn_test.get('status', 'Not tested')}")
    
    if diagnostics.get("recommendations"):
        print("\nRECOMMENDATIONS:")
        print("-" * 20)
        for i, rec in enumerate(diagnostics["recommendations"], 1):
            print(f"{i}. {rec}")