"""Export service for QuickBooks Auto Reporter.

Handles CSV and Excel export functionality with professional formatting.
"""

import csv
import datetime as dt
import json
import os
from io import StringIO
from typing import List, Dict, Any, Optional

from ..config import REPORT_CONFIGS, get_file_paths
from ..utils.file_utils import compute_data_hash, should_create_snapshot, save_hash, create_snapshot
from ..utils.logging_utils import log_success, log_error, log_data


def render_csv(headers: List[str], rows: List[List[str]]) -> str:
    """Render data as CSV string.
    
    Args:
        headers: List of column headers
        rows: List of data rows
        
    Returns:
        CSV content as string
    """
    sio = StringIO()
    w = csv.writer(sio, lineterminator="\n")
    w.writerow(headers)
    for r in rows:
        w.writerow([c if c is not None else "" for c in r])
    return sio.getvalue()


def export_to_csv(headers: List[str], rows: List[List[str]], out_dir: str, report_key: str) -> bool:
    """Export data to CSV file.
    
    Args:
        headers: List of column headers
        rows: List of data rows
        out_dir: Output directory path
        report_key: Report configuration key
        
    Returns:
        True if successful, False otherwise
    """
    try:
        file_paths = get_file_paths(out_dir, report_key)
        csv_path = file_paths["main_csv"]
        
        csv_content = render_csv(headers, rows)
        
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            f.write(csv_content)
        
        log_success(f"CSV file created: {os.path.basename(csv_path)}", out_dir)
        return True
        
    except Exception as e:
        log_error(f"CSV export failed for {report_key}: {e}", out_dir)
        return False


def export_to_excel(headers: List[str], rows: List[List[str]], out_dir: str, report_key: str) -> bool:
    """Export data to Excel file with professional formatting.
    
    Args:
        headers: List of column headers
        rows: List of data rows
        out_dir: Output directory path
        report_key: Report configuration key
        
    Returns:
        True if successful, False otherwise
    """
    try:
        file_paths = get_file_paths(out_dir, report_key)
        excel_path = file_paths["excel_file"]
        
        # Try using Excel MCP first for enhanced functionality
        try:
            return _create_excel_with_mcp(headers, rows, excel_path, report_key, out_dir)
        except Exception as mcp_error:
            log_error(f"Excel MCP failed, falling back to openpyxl: {mcp_error}", out_dir)
            return _create_excel_with_openpyxl(headers, rows, excel_path, report_key, out_dir)
            
    except Exception as e:
        log_error(f"Excel export failed for {report_key}: {e}", out_dir)
        return False


def _create_excel_with_mcp(headers: List[str], rows: List[List[str]], excel_path: str, report_key: str, out_dir: str) -> bool:
    """Create Excel file using Excel MCP with enhanced styling.
    
    Args:
        headers: List of column headers
        rows: List of data rows
        excel_path: Path for Excel file
        report_key: Report configuration key
        out_dir: Output directory path
        
    Returns:
        True if successful, False otherwise
    """
    # This is a placeholder for Excel MCP integration
    # In a real implementation, this would use Excel MCP functions
    # For now, we'll fall back to openpyxl
    return _create_excel_with_openpyxl(headers, rows, excel_path, report_key, out_dir)


def _create_excel_with_openpyxl(headers: List[str], rows: List[List[str]], excel_path: str, report_key: str, out_dir: str) -> bool:
    """Create Excel file using openpyxl with professional formatting.
    
    Args:
        headers: List of column headers
        rows: List of data rows
        excel_path: Path for Excel file
        report_key: Report configuration key
        out_dir: Output directory path
        
    Returns:
        True if successful, False otherwise
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = REPORT_CONFIGS[report_key]["name"][:31]  # Excel sheet name limit

        # Enhanced styling
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Border styles
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=str(header))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows with alternating colors
        data_fill_1 = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )
        data_fill_2 = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )

        for row_idx, row_data in enumerate(rows, 2):
            fill = data_fill_1 if row_idx % 2 == 0 else data_fill_2
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=str(value) if value is not None else "",
                )
                cell.fill = fill
                cell.border = thin_border

                # Format numbers if they look like currency or numbers
                if isinstance(value, (int, float)):
                    cell.number_format = (
                        "#,##0.00" if isinstance(value, float) else "#,##0"
                    )

        # Auto-adjust column widths with better logic
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for cell in column:
                try:
                    cell_value = (
                        str(cell.value) if cell.value is not None else ""
                    )
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    pass

            # Set width with reasonable limits
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze the header row
        ws.freeze_panes = "A2"

        # Add filter to header row
        if headers:
            ws.auto_filter.ref = (
                f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
            )

        # Save the workbook
        wb.save(excel_path)
        log_success(f"Excel file created with openpyxl: {os.path.basename(excel_path)}", out_dir)
        return True

    except ImportError as ie:
        log_error(f"Excel creation failed: openpyxl not available for {report_key}. Error: {ie}", out_dir)
        return False
    except Exception as excel_error:
        log_error(f"Excel creation error for {report_key}: {excel_error}", out_dir)
        return False


def handle_change_detection(csv_content: str, out_dir: str, report_key: str) -> Dict[str, Any]:
    """Handle change detection and snapshot creation.
    
    Args:
        csv_content: CSV content as string
        out_dir: Output directory path
        report_key: Report configuration key
        
    Returns:
        Dictionary with change detection results
    """
    new_hash = compute_data_hash(csv_content)
    changed = should_create_snapshot(out_dir, report_key, new_hash)
    
    result = {
        "hash": new_hash,
        "changed": changed,
        "snapshot_created": False,
        "snapshot_path": None
    }
    
    if changed:
        snapshot_path = create_snapshot(out_dir, report_key, csv_content)
        if snapshot_path:
            save_hash(out_dir, report_key, new_hash)
            result["snapshot_created"] = True
            result["snapshot_path"] = snapshot_path
            log_success(f"Snapshot created: {os.path.basename(snapshot_path)}", out_dir)
    
    return result


def export_report_with_change_detection(
    headers: List[str],
    rows: List[List[str]],
    out_dir: str,
    report_key: str,
    create_excel: bool = True
) -> Dict[str, Any]:
    """Export report with change detection and multiple format support.
    
    Args:
        headers: List of column headers
        rows: List of data rows
        out_dir: Output directory path
        report_key: Report configuration key
        create_excel: Whether to create Excel file
        
    Returns:
        Dictionary with export results
    """
    result = {
        "rows": len(rows),
        "excel_created": False,
        "csv_created": False,
        "changed": False,
        "snapshot_created": False,
        "timestamp": dt.datetime.now(),
        "report_name": REPORT_CONFIGS[report_key]["name"]
    }
    
    try:
        # Generate CSV content
        csv_content = render_csv(headers, rows)
        
        # Handle change detection
        change_result = handle_change_detection(csv_content, out_dir, report_key)
        result.update(change_result)
        
        # Always write the main CSV file
        file_paths = get_file_paths(out_dir, report_key)
        with open(file_paths["main_csv"], "w", newline="", encoding="utf-8") as f:
            f.write(csv_content)
        result["csv_created"] = True
        
        # Create Excel file if requested
        if create_excel:
            result["excel_created"] = export_to_excel(headers, rows, out_dir, report_key)
        
        # Log result
        if result["changed"]:
            log_success(
                f"{result['report_name']} changed. "
                f"CSV: ✅, Excel: {'✅' if result['excel_created'] else '❌'}, "
                f"Snapshot: {'✅' if result['snapshot_created'] else '❌'} "
                f"(rows={result['rows']})",
                out_dir
            )
        else:
            log_data(
                f"{result['report_name']} no change. "
                f"CSV: ✅, Excel: {'✅' if result['excel_created'] else '❌'} "
                f"(rows={result['rows']})",
                out_dir
            )
        
        return result
        
    except Exception as e:
        log_error(f"Export failed for {report_key}: {e}", out_dir)
        result["error"] = str(e)
        return result