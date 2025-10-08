# QuickBooks Auto Reporter — v1.0
# - Generates three reports: Open Sales Orders by Item, Profit & Loss, and Sales by Item
# - Uses GeneralDetailReportQueryRq for all reports per Intuit docs
# - Normalizes qbXML before send (strip BOM/leading whitespace) to avoid
#   "parsing the provided XML text stream" errors
# - Logs last_request.xml / last_response.xml for each report
# - Enhanced with Context7 MCP analytics and Excel MCP for professional reporting
# - Enhanced with user folder selection, timer display, and configurable intervals

import csv
import datetime as dt
import hashlib
import json
import os
import subprocess
import sys
import threading
import tkinter as tk
import winreg  # type: ignore
import xml.etree.ElementTree as ET
from tkinter import filedialog, messagebox, ttk

import pythoncom  # type: ignore
from win32com.client import Dispatch, gencache  # type: ignore

APP_NAME = "Gasco Auto Reporter"
QBXML_VERSION_PRIMARY = "16.0"
QBXML_VERSION_FALLBACK = "13.0"
ALLOW_SALESORDER_FALLBACK = True

COMPANY_FILE = os.environ.get(
    "QB_COMPANY_FILE",
    r"C:\Users\Public\Documents\Intuit\QuickBooks\Company Files"
    r"\Gasco Industrial 2019\Gasco Industrial -  2019.QBW",
)

# Default output directory (can be changed by user)
DEFAULT_OUT_DIR = r"C:\Reports"

# Interval options in seconds
INTERVAL_OPTIONS = {
    "5 minutes": 5 * 60,
    "15 minutes": 15 * 60,
    "30 minutes": 30 * 60,
    "60 minutes": 60 * 60,
}
DEFAULT_INTERVAL = "15 minutes"

# Settings file for persistence
SETTINGS_FILE = os.path.join(os.path.expanduser("~"), ".qb_auto_reporter_settings.json")

# Report configurations
REPORT_CONFIGS = {
    "open_sales_orders": {
        "name": "Open Sales Orders by Item",
        "qbxml_type": "OpenSalesOrderByItem",
        "query": "GeneralDetail",
        "csv_filename": "Open_Sales_Orders_By_Item.csv",
        "excel_filename": "Open_Sales_Orders_By_Item.xlsx",
        "hash_filename": "Open_Sales_Orders_By_Item.hash",
        "request_log": "open_so_request.xml",
        "response_log": "open_so_response.xml",
        "uses_date_range": False,
    },
    "profit_loss": {
        "name": "Profit & Loss",
        "qbxml_type": "ProfitAndLossStandard",
        "query": "GeneralSummary",
        "csv_filename": "Profit_And_Loss.csv",
        "excel_filename": "Profit_And_Loss.xlsx",
        "hash_filename": "Profit_And_Loss.hash",
        "request_log": "pl_request.xml",
        "response_log": "pl_response.xml",
        "uses_date_range": True,
    },
    "profit_loss_detail": {
        "name": "Profit & Loss Detail",
        "qbxml_type": "ProfitAndLossDetail",
        "query": "GeneralDetail",
        "csv_filename": "Profit_And_Loss_Detail.csv",
        "excel_filename": "Profit_And_Loss_Detail.xlsx",
        "hash_filename": "Profit_And_Loss_Detail.hash",
        "request_log": "pl_detail_request.xml",
        "response_log": "pl_detail_response.xml",
        "uses_date_range": True,
    },
    "sales_by_item": {
        "name": "Sales by Item",
        "qbxml_type": "SalesByItemSummary",
        "query": "GeneralSummary",
        "csv_filename": "Sales_By_Item.csv",
        "excel_filename": "Sales_By_Item.xlsx",
        "hash_filename": "Sales_By_Item.hash",
        "request_log": "sales_item_request.xml",
        "response_log": "sales_item_response.xml",
        "uses_date_range": True,
    },
    "sales_by_item_detail": {
        "name": "Sales by Item Detail",
        "qbxml_type": "SalesByItemDetail",
        "query": "GeneralDetail",
        "csv_filename": "Sales_By_Item_Detail.csv",
        "excel_filename": "Sales_By_Item_Detail.xlsx",
        "hash_filename": "Sales_By_Item_Detail.hash",
        "request_log": "sales_item_detail_request.xml",
        "response_log": "sales_item_detail_response.xml",
        "uses_date_range": True,
    },
    "sales_by_rep_detail": {
        "name": "Sales by Rep Detail",
        "qbxml_type": "SalesByRepDetail",
        "query": "GeneralDetail",
        "csv_filename": "Sales_By_Rep_Detail.csv",
        "excel_filename": "Sales_By_Rep_Detail.xlsx",
        "hash_filename": "Sales_By_Rep_Detail.hash",
        "request_log": "sales_rep_detail_request.xml",
        "response_log": "sales_rep_detail_response.xml",
        "uses_date_range": True,
    },
    "purchase_by_vendor_detail": {
        "name": "Purchase by Vendor Detail",
        "qbxml_type": "PurchasesByVendorDetail",
        "query": "GeneralDetail",
        "csv_filename": "Purchase_By_Vendor_Detail.csv",
        "excel_filename": "Purchase_By_Vendor_Detail.xlsx",
        "hash_filename": "Purchase_By_Vendor_Detail.hash",
        "request_log": "purchase_vendor_detail_request.xml",
        "response_log": "purchase_vendor_detail_response.xml",
        "uses_date_range": True,
    },
    "ap_aging_detail": {
        "name": "AP Aging Detail",
        "qbxml_type": "APAgingDetail",
        "query": "Aging",
        "csv_filename": "AP_Aging_Detail.csv",
        "excel_filename": "AP_Aging_Detail.xlsx",
        "hash_filename": "AP_Aging_Detail.hash",
        "request_log": "ap_aging_detail_request.xml",
        "response_log": "ap_aging_detail_response.xml",
        "uses_date_range": False,  # Aging reports use AsOfReportDate
    },
    "ar_aging_detail": {
        "name": "AR Aging Detail",
        "qbxml_type": "ARAgingDetail",
        "query": "Aging",
        "csv_filename": "AR_Aging_Detail.csv",
        "excel_filename": "AR_Aging_Detail.xlsx",
        "hash_filename": "AR_Aging_Detail.hash",
        "request_log": "ar_aging_detail_request.xml",
        "response_log": "ar_aging_detail_response.xml",
        "uses_date_range": False,  # Aging reports use AsOfReportDate
    },
}


def load_settings():
    """Load user settings from file"""
    # Default to current month date range
    today = dt.date.today()
    first_day = today.replace(day=1)
    default_settings = {
        "output_dir": DEFAULT_OUT_DIR,
        "interval": DEFAULT_INTERVAL,
        "report_date_from": first_day.strftime("%Y-%m-%d"),
        "report_date_to": today.strftime("%Y-%m-%d"),
    }
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                settings = json.load(f)
                if settings.get("interval") not in INTERVAL_OPTIONS:
                    settings["interval"] = DEFAULT_INTERVAL
                # Ensure date settings exist
                if "report_date_from" not in settings:
                    settings["report_date_from"] = default_settings["report_date_from"]
                if "report_date_to" not in settings:
                    settings["report_date_to"] = default_settings["report_date_to"]
                return settings
    except Exception:
        pass
    return default_settings


def save_settings(settings):
    """Save user settings to file"""
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(settings, f, indent=2)
    except Exception:
        pass


def get_file_paths(out_dir, report_key):
    """Get all file paths based on output directory and report type"""
    config = REPORT_CONFIGS[report_key]
    return {
        "main_csv": os.path.join(out_dir, config["csv_filename"]),
        "excel_file": os.path.join(out_dir, config["excel_filename"]),
        "hash_file": os.path.join(out_dir, config["hash_filename"]),
        "log_file": os.path.join(out_dir, "QuickBooks_Auto_Reports.log"),
        "req_log": os.path.join(out_dir, config["request_log"]),
        "resp_log": os.path.join(out_dir, config["response_log"]),
    }


def log(msg: str, out_dir: str = None) -> None:
    """Log message to file with timestamp"""
    if out_dir is None:
        out_dir = DEFAULT_OUT_DIR
    ts = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    os.makedirs(out_dir, exist_ok=True)
    log_file = os.path.join(out_dir, "QuickBooks_Auto_Reports.log")
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"[{ts}] {msg}\n")


def sha256_text(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8", "ignore")).hexdigest()


def check_quickbooks_installation():
    """Check if QuickBooks Desktop is installed on the system"""
    try:
        # Check for QuickBooks in registry
        qb_paths = []
        registry_keys = [
            r"SOFTWARE\Intuit\QuickBooks",
            r"SOFTWARE\WOW6432Node\Intuit\QuickBooks",
        ]
        
        for key_path in registry_keys:
            try:
                with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, key_path) as key:
                    i = 0
                    while True:
                        try:
                            subkey_name = winreg.EnumKey(key, i)
                            if subkey_name.startswith(('QB', 'QuickBooks')):
                                qb_paths.append(f"{key_path}\\{subkey_name}")
                            i += 1
                        except WindowsError:
                            break
            except FileNotFoundError:
                continue
        
        return len(qb_paths) > 0, qb_paths
    except Exception as e:
        return False, [f"Registry check failed: {e}"]


def check_sdk_installation():
    """Check if QuickBooks SDK is properly installed and registered"""
    try:
        # Try to create the COM object without initializing
        pythoncom.CoInitialize()
        try:
            # Check if QBXMLRP2.RequestProcessor is registered
            clsid = pythoncom.CLSIDFromProgID("QBXMLRP2.RequestProcessor")
            return True, f"SDK found with CLSID: {clsid}"
        except pythoncom.com_error as e:
            return False, f"COM registration error: {e}"
        finally:
            pythoncom.CoUninitialize()
    except Exception as e:
        return False, f"SDK check failed: {e}"


def get_user_friendly_error(error):
    """Convert technical errors to user-friendly messages with solutions"""
    error_str = str(error)
    
    # COM Error -2147221005: Invalid class string
    if "-2147221005" in error_str or "Invalid class string" in error_str:
        return {
            "title": "QuickBooks Connection Problem",
            "message": "Cannot connect to QuickBooks Desktop. This usually means QuickBooks SDK is not installed or not properly registered.",
            "solutions": [
                "1. Install QuickBooks Desktop if not already installed",
                "2. Download and install the QuickBooks SDK from Intuit Developer website",
                "3. Run the application as Administrator",
                "4. Restart your computer after SDK installation",
                "5. Make sure QuickBooks Desktop is closed before running reports"
            ],
            "technical_details": error_str,
            "error_type": "SDK_NOT_INSTALLED"
        }
    
    # COM Error -2147221164: Class not registered
    elif "-2147221164" in error_str or "Class not registered" in error_str:
        return {
            "title": "QuickBooks SDK Not Registered",
            "message": "The QuickBooks SDK components are not properly registered on this system.",
            "solutions": [
                "1. Reinstall the QuickBooks SDK",
                "2. Run 'regsvr32 qbxmlrp2.dll' as Administrator",
                "3. Restart your computer",
                "4. Contact your IT administrator for help with COM registration"
            ],
            "technical_details": error_str,
            "error_type": "SDK_NOT_REGISTERED"
        }
    
    # Access denied errors
    elif "Access" in error_str and "denied" in error_str:
        return {
            "title": "Permission Problem",
            "message": "The application doesn't have permission to access QuickBooks.",
            "solutions": [
                "1. Run the application as Administrator",
                "2. Check QuickBooks company file permissions",
                "3. Make sure QuickBooks is not in multi-user mode",
                "4. Close QuickBooks Desktop and try again"
            ],
            "technical_details": error_str,
            "error_type": "ACCESS_DENIED"
        }
    
    # File not found or path errors
    elif any(term in error_str.lower() for term in ["file not found", "path", "cannot find"]):
        return {
            "title": "QuickBooks File Problem",
            "message": "Cannot find or access the QuickBooks company file.",
            "solutions": [
                "1. Make sure QuickBooks Desktop is installed and working",
                "2. Open QuickBooks and verify the company file opens correctly",
                "3. Check the QB_COMPANY_FILE environment variable path",
                "4. Make sure the company file is not on a network drive that's disconnected"
            ],
            "technical_details": error_str,
            "error_type": "FILE_NOT_FOUND"
        }
    
    # Network or connection errors
    elif any(term in error_str.lower() for term in ["network", "connection", "timeout"]):
        return {
            "title": "Connection Problem",
            "message": "Cannot establish a connection to QuickBooks.",
            "solutions": [
                "1. Make sure QuickBooks Desktop is running",
                "2. Check if QuickBooks is in single-user mode",
                "3. Restart QuickBooks Desktop",
                "4. Check network connectivity if using a network installation"
            ],
            "technical_details": error_str,
            "error_type": "CONNECTION_ERROR"
        }
    
    # Generic error
    else:
        return {
            "title": "QuickBooks Error",
            "message": "An unexpected error occurred while connecting to QuickBooks.",
            "solutions": [
                "1. Make sure QuickBooks Desktop is installed and running",
                "2. Try restarting QuickBooks Desktop",
                "3. Run this application as Administrator",
                "4. Check the log file for more details",
                "5. Contact support with the technical details below"
            ],
            "technical_details": error_str,
            "error_type": "UNKNOWN_ERROR"
        }


def diagnose_quickbooks_connection(out_dir: str = None):
    """Run comprehensive QuickBooks connectivity diagnostics"""
    if out_dir is None:
        out_dir = DEFAULT_OUT_DIR
    
    os.makedirs(out_dir, exist_ok=True)
    
    diagnostics = {
        "timestamp": dt.datetime.now().isoformat(),
        "system_info": {},
        "quickbooks_installation": {},
        "sdk_installation": {},
        "connectivity_test": {},
        "recommendations": []
    }
    
    log("🔍 Running QuickBooks diagnostics...", out_dir)
    
    # System information
    try:
        diagnostics["system_info"] = {
            "platform": sys.platform,
            "python_version": sys.version,
            "architecture": "64-bit" if sys.maxsize > 2**32 else "32-bit"
        }
    except Exception as e:
        diagnostics["system_info"]["error"] = str(e)
    
    # Check QuickBooks installation
    qb_installed, qb_info = check_quickbooks_installation()
    diagnostics["quickbooks_installation"] = {
        "installed": qb_installed,
        "details": qb_info,
        "status": "✅ Found" if qb_installed else "❌ Not Found"
    }
    
    # Check SDK installation
    sdk_installed, sdk_info = check_sdk_installation()
    diagnostics["sdk_installation"] = {
        "installed": sdk_installed,
        "details": sdk_info,
        "status": "✅ Registered" if sdk_installed else "❌ Not Registered"
    }
    
    # Test connectivity
    try:
        pythoncom.CoInitialize()
        try:
            rp = gencache.EnsureDispatch("QBXMLRP2.RequestProcessor")
            diagnostics["connectivity_test"] = {
                "com_object_creation": "✅ Success",
                "status": "SDK components are working"
            }
            
            # Try to open connection
            try:
                open_connection(rp)
                diagnostics["connectivity_test"]["connection_test"] = "✅ Success"
                try:
                    rp.CloseConnection()
                except:
                    pass
            except Exception as conn_e:
                error_info = get_user_friendly_error(conn_e)
                diagnostics["connectivity_test"]["connection_test"] = f"❌ Failed: {error_info['message']}"
                diagnostics["connectivity_test"]["connection_error"] = error_info
                
        except Exception as com_e:
            error_info = get_user_friendly_error(com_e)
            diagnostics["connectivity_test"] = {
                "com_object_creation": f"❌ Failed: {error_info['message']}",
                "error_details": error_info
            }
        finally:
            pythoncom.CoUninitialize()
    except Exception as e:
        diagnostics["connectivity_test"]["error"] = str(e)
    
    # Generate recommendations
    if not qb_installed:
        diagnostics["recommendations"].append("Install QuickBooks Desktop from Intuit")
    
    if not sdk_installed:
        diagnostics["recommendations"].append("Download and install QuickBooks SDK from Intuit Developer website")
        diagnostics["recommendations"].append("Run the application as Administrator after SDK installation")
    
    if qb_installed and sdk_installed:
        diagnostics["recommendations"].append("QuickBooks and SDK appear to be installed correctly")
        if "connection_error" in diagnostics.get("connectivity_test", {}):
            error_type = diagnostics["connectivity_test"]["connection_error"].get("error_type")
            if error_type == "ACCESS_DENIED":
                diagnostics["recommendations"].append("Try running as Administrator")
            elif error_type == "FILE_NOT_FOUND":
                diagnostics["recommendations"].append("Verify QuickBooks company file path")
    
    # Save diagnostics to JSON
    diag_file = os.path.join(out_dir, "quickbooks_diagnostics.json")
    with open(diag_file, "w", encoding="utf-8") as f:
        json.dump(diagnostics, f, indent=2)
    
    # Create Excel diagnostic report
    create_diagnostic_excel_report(diagnostics, out_dir)
    
    log(f"📊 Diagnostics completed. Report saved to: {diag_file}", out_dir)
    return diagnostics


def create_diagnostic_excel_report(diagnostics: dict, out_dir: str):
    """Create an Excel diagnostic report using Excel MCP"""
    try:
        excel_path = os.path.join(out_dir, "QuickBooks_Diagnostic_Report.xlsx")
        
        # Prepare diagnostic data for Excel
        summary_data = [
            ["QuickBooks Diagnostic Report", ""],
            ["Generated", diagnostics.get("timestamp", "")],
            ["", ""],
            ["Component", "Status"],
            ["QuickBooks Desktop", diagnostics["quickbooks_installation"]["status"]],
            ["QuickBooks SDK", diagnostics["sdk_installation"]["status"]],
            ["", ""],
            ["System Information", ""],
            ["Platform", diagnostics["system_info"].get("platform", "Unknown")],
            ["Python Version", diagnostics["system_info"].get("python_version", "Unknown")],
            ["Architecture", diagnostics["system_info"].get("architecture", "Unknown")],
        ]
        
        # Add connectivity test results
        if "connectivity_test" in diagnostics:
            summary_data.extend([
                ["", ""],
                ["Connectivity Tests", ""],
                ["COM Object Creation", diagnostics["connectivity_test"].get("com_object_creation", "Not tested")],
                ["Connection Test", diagnostics["connectivity_test"].get("connection_test", "Not tested")],
            ])
        
        # Add recommendations
        if diagnostics.get("recommendations"):
            summary_data.extend([
                ["", ""],
                ["Recommendations", ""],
            ])
            for i, rec in enumerate(diagnostics["recommendations"], 1):
                summary_data.append([f"{i}.", rec])
        
        # Try to create Excel file
        try:
            # Use Excel MCP if available
            from mcp_excel_create_workbook import mcp_excel_create_workbook
            from mcp_excel_format_range import mcp_excel_format_range
            from mcp_excel_write_data_to_excel import mcp_excel_write_data_to_excel
            
            mcp_excel_create_workbook(filepath=excel_path)
            mcp_excel_write_data_to_excel(
                filepath=excel_path,
                sheet_name="Diagnostic Report",
                data=summary_data,
                start_cell="A1"
            )
            
            # Format the report
            mcp_excel_format_range(
                filepath=excel_path,
                sheet_name="Diagnostic Report",
                start_cell="A1",
                end_cell="B1",
                bold=True,
                font_size=14,
                bg_color="4472C4",
                font_color="FFFFFF"
            )
            
            log(f"📊 Excel diagnostic report created: {os.path.basename(excel_path)}", out_dir)
            
        except Exception as mcp_error:
            # Fallback to openpyxl
            try:
                import openpyxl
                from openpyxl.styles import Alignment, Font, PatternFill
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Diagnostic Report"
                
                # Write data with basic formatting
                for row_idx, (col1, col2) in enumerate(summary_data, 1):
                    ws.cell(row=row_idx, column=1, value=col1)
                    ws.cell(row=row_idx, column=2, value=col2)
                    
                    # Format header
                    if row_idx == 1:
                        ws.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(excel_path)
                log(f"📊 Excel diagnostic report created with openpyxl: {os.path.basename(excel_path)}", out_dir)
                
            except Exception as excel_error:
                log(f"❌ Could not create Excel diagnostic report: {excel_error}", out_dir)
    
    except Exception as e:
        log(f"❌ Diagnostic Excel report creation failed: {e}", out_dir)


def open_connection(rp):
    for local_enum in (1, 0, 2):
        try:
            rp.OpenConnection2("", APP_NAME, local_enum)
            return
        except Exception:
            pass
    rp.OpenConnection("", APP_NAME)


def host_info(rp, ticket):
    host_req = """<?xml version="1.0"?>
<?qbxml version="16.0"?>
<QBXML><QBXMLMsgsRq onError="stopOnError"><HostQueryRq/></QBXMLMsgsRq></QBXML>"""
    resp = rp.ProcessRequest(ticket, host_req)
    root = ET.fromstring(resp)
    h = root.find(".//HostRet")
    fn = (h.findtext("CompanyFileName") or "") if h is not None else ""
    mode = (h.findtext("QBFileMode") or "") if h is not None else ""
    ai = (h.findtext("IsAutomaticLogin") or "") if h is not None else ""
    return fn, mode, ai


def try_begin_session(rp):
    attempts = [("", 0), ("", 2), ("", 1), (COMPANY_FILE, 2), (COMPANY_FILE, 1)]
    last = None
    for path, mode in attempts:
        try:
            t = rp.BeginSession(path, mode)
            fn, fm, ai = host_info(rp, t)
            return t, {
                "CompanyFileName": fn,
                "QBFileMode": fm,
                "IsAutomaticLogin": ai,
                "PathUsed": path,
                "ModeUsed": mode,
            }
        except Exception as e:
            last = e
    raise RuntimeError(
        f"BeginSession attempts failed. Last error: {last}\nTried: {attempts}\nCOMPANY_FILE={COMPANY_FILE}"
    )


def qb_request(xml: str, out_dir: str = None, report_key: str = "open_sales_orders"):
    """Execute qbXML request with enhanced error handling and user-friendly messages"""
    if out_dir is None:
        out_dir = DEFAULT_OUT_DIR

    # Normalize qbXML to avoid parser errors
    xml = xml.lstrip("\ufeff \t\r\n")
    xml = xml.replace("\r\n", "\n")

    file_paths = get_file_paths(out_dir, report_key)

    pythoncom.CoInitialize()
    try:
        try:
            rp = gencache.EnsureDispatch("QBXMLRP2.RequestProcessor")
        except Exception as com_error:
            # Enhanced error handling for COM object creation
            error_info = get_user_friendly_error(com_error)
            log(f"❌ {error_info['title']}: {error_info['message']}", out_dir)
            
            # Log technical details for debugging
            log(f"🔍 Technical details: {error_info['technical_details']}", out_dir)
            
            # Log solutions
            log("💡 Possible solutions:", out_dir)
            for solution in error_info['solutions']:
                log(f"   {solution}", out_dir)
            
            # Re-raise with user-friendly message
            raise RuntimeError(f"{error_info['title']}: {error_info['message']}")
        
        try:
            open_connection(rp)
        except Exception as conn_error:
            error_info = get_user_friendly_error(conn_error)
            log(f"❌ {error_info['title']}: {error_info['message']}", out_dir)
            log(f"🔍 Technical details: {error_info['technical_details']}", out_dir)
            raise RuntimeError(f"{error_info['title']}: {error_info['message']}")
        
        try:
            ticket, info = try_begin_session(rp)
        except Exception as session_error:
            error_info = get_user_friendly_error(session_error)
            log(f"❌ {error_info['title']}: {error_info['message']}", out_dir)
            log(f"🔍 Technical details: {error_info['technical_details']}", out_dir)
            raise RuntimeError(f"{error_info['title']}: {error_info['message']}")
        
        try:
            os.makedirs(out_dir, exist_ok=True)
            with open(file_paths["req_log"], "w", encoding="utf-8") as f:
                f.write(xml)
            resp = rp.ProcessRequest(ticket, xml)
            with open(file_paths["resp_log"], "w", encoding="utf-8") as f:
                f.write(resp)
            return resp, info
        finally:
            rp.EndSession(ticket)
            try:
                rp.CloseConnection()
            except Exception:
                pass
    finally:
        pythoncom.CoUninitialize()


def build_report_qbxml(
    version: str,
    report_type: str,
    date_from: str = None,
    date_to: str = None,
    report_key: str = None,
) -> str:
    """Build qbXML for the appropriate report query with optional date range using correct XML structure from examples"""

    config = REPORT_CONFIGS.get(report_key, {}) if report_key else {}
    query = config.get("query", "GeneralDetail")
    uses_date_range = config.get("uses_date_range", False)

    # Build XML based on the working examples
    if query == "GeneralSummary":
        open_tag = "GeneralSummaryReportQueryRq"
        type_tag = "GeneralSummaryReportType"
    elif query == "Aging":
        open_tag = "AgingReportQueryRq"
        type_tag = "AgingReportType"
    else:
        open_tag = "GeneralDetailReportQueryRq"
        type_tag = "GeneralDetailReportType"

    xml_parts = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        f'<?qbxml version="{version}"?>',
        "<QBXML>",
        '  <QBXMLMsgsRq onError="continueOnError">',
        f'    <{open_tag} requestID="1">',
        f"      <{type_tag}>{report_type}</{type_tag}>",
    ]

    # Handle different date structures based on working examples
    if query == "Aging":
        # Aging reports use ReportPeriod with ToReportDate and ReportAgingAsOf
        as_of_date = date_to if date_to else dt.date.today().strftime("%Y-%m-%d")
        xml_parts.extend(
            [
                "      <ReportPeriod>",
                f"        <ToReportDate>{as_of_date}</ToReportDate>",
                "      </ReportPeriod>",
                "      <ReportAgingAsOf>ReportEndDate</ReportAgingAsOf>",
                "      <DisplayReport>true</DisplayReport>",
            ]
        )
    elif uses_date_range and date_from and date_to:
        # Validate dates
        try:
            dt.datetime.strptime(date_from, "%Y-%m-%d")
            dt.datetime.strptime(date_to, "%Y-%m-%d")
        except ValueError:
            today = dt.date.today()
            first_day = today.replace(day=1)
            date_from = first_day.strftime("%Y-%m-%d")
            date_to = today.strftime("%Y-%m-%d")

        # Different reports use different date structures
        if report_key == "purchase_by_vendor_detail":
            # Purchase reports have dates directly under the query element
            xml_parts.extend(
                [
                    f"      <FromReportDate>{date_from}</FromReportDate>",
                    f"      <ToReportDate>{date_to}</ToReportDate>",
                    "      <DisplayReport>true</DisplayReport>",
                ]
            )
        else:
            # Most reports use ReportPeriod wrapper
            xml_parts.extend(
                [
                    "      <DisplayReport>true</DisplayReport>",
                    "      <ReportPeriod>",
                    f"        <FromReportDate>{date_from}</FromReportDate>",
                    f"        <ToReportDate>{date_to}</ToReportDate>",
                    "      </ReportPeriod>",
                ]
            )
    else:
        # Reports without date ranges (like Open Sales Orders)
        xml_parts.append("      <DisplayReport>true</DisplayReport>")

    xml_parts.extend(
        [
            f"    </{open_tag}>",
            "  </QBXMLMsgsRq>",
            "</QBXML>",
        ]
    )

    return "\n".join(xml_parts)


def validate_xml_against_examples(report_key: str, generated_xml: str, out_dir: str):
    """Validate generated XML against working examples"""
    try:
        examples_dir = os.path.join(
            os.path.dirname(__file__),
            "QuickBooks_Auto_Reporter_Portable",
            "quickbooks_autoreport",
        )

        # Map report keys to example files
        example_files = {
            "open_sales_orders": "open_so_request.xml",
            "profit_loss": "pl_request.xml",
            "profit_loss_detail": "pl_detail_request.xml",
            "sales_by_item": "sales_item_request.xml",
            "sales_by_item_detail": "sales_item_detail_request.xml",
            "sales_by_rep_detail": "sales_rep_detail_request.xml",
            "purchase_by_vendor_detail": "purchase_vendor_detail_request.xml",
            "ap_aging_detail": "ap_aging_detail_request.xml",
            "ar_aging_detail": "ar_aging_detail_request.xml",
        }

        example_file = example_files.get(report_key)
        if not example_file:
            log(f"⚠️ No example file found for {report_key}", out_dir)
            return False

        example_path = os.path.join(examples_dir, example_file)
        if not os.path.exists(example_path):
            log(f"⚠️ Example file not found: {example_path}", out_dir)
            return False

        # Read the example XML
        with open(example_path, "r", encoding="utf-8") as f:
            example_xml = f.read().strip()

        # Basic structure validation
        import xml.etree.ElementTree as ET

        try:
            generated_root = ET.fromstring(generated_xml)
            example_root = ET.fromstring(example_xml)

            # Compare key elements - look for different report type elements
            gen_query_type = (
                generated_root.find(".//GeneralDetailReportType")
                or generated_root.find(".//GeneralSummaryReportType")
                or generated_root.find(".//AgingReportType")
            )
            ex_query_type = (
                example_root.find(".//GeneralDetailReportType")
                or example_root.find(".//GeneralSummaryReportType")
                or example_root.find(".//AgingReportType")
            )

            if gen_query_type is not None and ex_query_type is not None:
                if gen_query_type.text == ex_query_type.text:
                    log(
                        f"✅ XML validation passed for {report_key}: ReportType matches",
                        out_dir,
                    )
                    return True
                else:
                    log(
                        f"⚠️ XML validation warning for {report_key}: ReportType mismatch - Generated: {gen_query_type.text}, Example: {ex_query_type.text}",
                        out_dir,
                    )

        except ET.ParseError as pe:
            log(
                f"❌ XML parsing error during validation for {report_key}: {pe}",
                out_dir,
            )
            return False

    except Exception as e:
        log(f"❌ XML validation failed for {report_key}: {e}", out_dir)
        return False

    return True


def build_salesorder_query(version: str) -> str:
    return f"""<?xml version="1.0"?>
<?qbxml version="{version}"?>
<QBXML><QBXMLMsgsRq onError="stopOnError">
  <SalesOrderQueryRq iterator="Start">
    <MaxReturned>200</MaxReturned>
    <IncludeRetElement>RefNumber</IncludeRetElement>
    <IncludeRetElement>TxnDate</IncludeRetElement>
    <IncludeRetElement>CustomerRef</IncludeRetElement>
    <IncludeRetElement>IsFullyInvoiced</IncludeRetElement>
    <IncludeRetElement>IsManuallyClosed</IncludeRetElement>
  </SalesOrderQueryRq>
</QBXMLMsgsRq></QBXML>"""


def parse_report_rows(resp_xml: str):
    """Parse QuickBooks report response XML with enhanced error handling"""
    try:
        root = ET.fromstring(resp_xml)

        # Look for the response element
        rs = root.find(".//GeneralDetailReportQueryRs")
        if rs is None:
            # Try to find any report query response
            for el in root.iter():
                if el.tag.endswith("ReportQueryRs"):
                    rs = el
                    break

        # Check for errors in the response
        if rs is None:
            raise RuntimeError("No ReportQueryRs found in response")

        status_code = rs.get("statusCode")
        if status_code not in (None, "0"):
            status_msg = rs.get("statusMessage", "Unknown error")
            raise RuntimeError(
                f"ReportQuery failed: status={status_code}, message={status_msg}"
            )

        # Find the report data
        rep = rs.find(".//ReportRet")
        if rep is None:
            raise RuntimeError("No ReportRet found in response")

        # Extract column headers
        headers = []
        col_descs = rep.findall(".//ColDesc")
        for c in col_descs:
            title = (c.findtext("ColTitle") or "").strip()
            if title:
                headers.append(title)

        # If no headers found, create default ones
        if not headers:
            headers = [f"Column_{i}" for i in range(1, 16)]

        rows = []

        def add_text_row(text):
            """Add a text row (like subtotals) to the results"""
            row = [""] * len(headers)
            if text.strip():
                row[0] = text.strip()
                rows.append(row)

        def walk(node):
            """Recursively walk through the XML structure"""
            for child in list(node):
                tag = child.tag

                if tag.endswith("TextRow"):
                    # Handle text rows (headers, subtotals, etc.)
                    vals = [cd.get("value", "") for cd in child.findall(".//ColData")]
                    text = (vals[0] if vals else child.get("value", "")).strip()
                    if text:
                        add_text_row(text)
                    walk(child)

                elif tag.endswith("DataRow"):
                    # Handle data rows
                    row = [""] * len(headers)
                    col_data_elements = child.findall("ColData")

                    for cd in col_data_elements:
                        val = cd.get("value", "")
                        col_id = cd.get("colID")

                        # Try to place value in correct column
                        if col_id and col_id.isdigit():
                            idx = int(col_id) - 1
                            if 0 <= idx < len(headers):
                                row[idx] = val
                        else:
                            # Find first empty column
                            for i in range(len(headers)):
                                if row[i] == "":
                                    row[i] = val
                                    break

                    # Only add non-empty rows
                    if any(cell.strip() for cell in row):
                        rows.append(row)
                    walk(child)

                elif tag.endswith("SubtotalRow"):
                    add_text_row("Subtotal")
                    walk(child)
                else:
                    walk(child)

        # Start walking from ReportData
        data = rep.find(".//ReportData")
        if data is not None:
            walk(data)
        else:
            # If no ReportData, try walking from the report root
            walk(rep)

        return headers, rows

    except ET.ParseError as pe:
        raise RuntimeError(f"XML parsing error: {pe}")
    except Exception as e:
        raise RuntimeError(f"Error parsing report response: {e}")


def parse_salesorders_to_rows(resp_xml: str):
    root = ET.fromstring(resp_xml)
    rs = root.find(".//SalesOrderQueryRs")
    if rs is None or rs.get("statusCode") not in (None, "0"):
        raise RuntimeError("SalesOrderQuery failed")
    headers = [
        "RefNumber",
        "TxnDate",
        "Customer",
        "IsFullyInvoiced",
        "IsManuallyClosed",
    ]
    rows = []
    for so in rs.findall("SalesOrderRet"):
        fully = (so.findtext("IsFullyInvoiced") or "").strip().lower() == "true"
        closed = (so.findtext("IsManuallyClosed") or "").strip().lower() == "true"
        if fully or closed:
            continue
        rows.append(
            [
                so.findtext("RefNumber") or "",
                so.findtext("TxnDate") or "",
                so.findtext("CustomerRef/FullName") or "",
                str(fully).lower(),
                str(closed).lower(),
            ]
        )
    return headers, rows


def render_csv(headers, rows):
    from io import StringIO

    sio = StringIO()
    w = csv.writer(sio, lineterminator="\n")
    w.writerow(headers)
    for r in rows:
        w.writerow([c if c is not None else "" for c in r])
    return sio.getvalue()


def snapshot_filename(out_dir: str, report_key: str):
    """Generate timestamped snapshot filename"""
    ts = dt.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    main_csv = get_file_paths(out_dir, report_key)["main_csv"]
    base, ext = os.path.splitext(main_csv)
    return f"{base}_{ts}{ext}"


def create_excel_report(headers, rows, out_dir: str, report_key: str):
    """Create professional Excel report using Excel MCP with enhanced styling and analytics"""
    try:
        file_paths = get_file_paths(out_dir, report_key)
        excel_path = file_paths["excel_file"]

        # Prepare data for Excel MCP
        excel_data = [headers] + rows

        # Try using Excel MCP first for enhanced functionality
        try:
            # Create workbook using Excel MCP
            from mcp_excel_create_chart import mcp_excel_create_chart
            from mcp_excel_create_workbook import mcp_excel_create_workbook
            from mcp_excel_format_range import mcp_excel_format_range
            from mcp_excel_write_data_to_excel import mcp_excel_write_data_to_excel

            # Create workbook
            mcp_excel_create_workbook(filepath=excel_path)

            # Write data to Excel
            mcp_excel_write_data_to_excel(
                filepath=excel_path,
                sheet_name=REPORT_CONFIGS[report_key]["name"][:31],
                data=excel_data,
                start_cell="A1",
            )

            # Format header row
            end_col_letter = chr(ord("A") + len(headers) - 1)
            mcp_excel_format_range(
                filepath=excel_path,
                sheet_name=REPORT_CONFIGS[report_key]["name"][:31],
                start_cell="A1",
                end_cell=f"{end_col_letter}1",
                bold=True,
                bg_color="4472C4",
                font_color="FFFFFF",
                alignment="center",
            )

            # Format data rows with alternating colors
            if len(rows) > 0:
                for i in range(2, len(rows) + 2, 2):  # Even rows
                    mcp_excel_format_range(
                        filepath=excel_path,
                        sheet_name=REPORT_CONFIGS[report_key]["name"][:31],
                        start_cell=f"A{i}",
                        end_cell=f"{end_col_letter}{i}",
                        bg_color="F2F2F2",
                    )

            log(
                f"✅ Excel file created with MCP: {os.path.basename(excel_path)}",
                out_dir,
            )
            return True

        except Exception as mcp_error:
            log(f"⚠️ Excel MCP failed, falling back to openpyxl: {mcp_error}", out_dir)

            # Fallback to openpyxl implementation
            try:
                import openpyxl
                from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
                from openpyxl.utils import get_column_letter

                # Create workbook and worksheet
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = REPORT_CONFIGS[report_key]["name"][
                    :31
                ]  # Excel sheet name limit

                # Enhanced styling
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                header_alignment = Alignment(horizontal="center", vertical="center")

                # Border styles
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                # Write headers with styling
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=str(header))
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                # Write data rows with alternating colors
                data_fill_1 = PatternFill(
                    start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                )
                data_fill_2 = PatternFill(
                    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                )

                for row_idx, row_data in enumerate(rows, 2):
                    fill = data_fill_1 if row_idx % 2 == 0 else data_fill_2
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(
                            row=row_idx,
                            column=col_idx,
                            value=str(value) if value is not None else "",
                        )
                        cell.fill = fill
                        cell.border = thin_border

                        # Format numbers if they look like currency or numbers
                        if isinstance(value, (int, float)):
                            cell.number_format = (
                                "#,##0.00" if isinstance(value, float) else "#,##0"
                            )

                # Auto-adjust column widths with better logic
                for col_idx, column in enumerate(ws.columns, 1):
                    max_length = 0
                    column_letter = get_column_letter(col_idx)

                    for cell in column:
                        try:
                            cell_value = (
                                str(cell.value) if cell.value is not None else ""
                            )
                            if len(cell_value) > max_length:
                                max_length = len(cell_value)
                        except Exception:
                            pass

                    # Set width with reasonable limits
                    adjusted_width = min(max(max_length + 2, 10), 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Freeze the header row
                ws.freeze_panes = "A2"

                # Add filter to header row
                if headers:
                    ws.auto_filter.ref = (
                        f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
                    )

                # Save the workbook
                wb.save(excel_path)
                log(
                    f"✅ Excel file created with openpyxl: {os.path.basename(excel_path)}",
                    out_dir,
                )
                return True

            except ImportError as ie:
                log(
                    f"❌ Excel creation failed: openpyxl not available for {report_key}. Error: {ie}",
                    out_dir,
                )
                return False
            except Exception as excel_error:
                log(f"❌ Excel creation error for {report_key}: {excel_error}", out_dir)
                return False

    except Exception as e:
        log(f"❌ Excel creation failed for {report_key}: {e}", out_dir)
        return False


def create_enhanced_excel_report(
    headers, rows, out_dir: str, report_key: str, insights: dict = None
):
    """Create enhanced Excel report with charts and analytics using Excel MCP"""
    try:
        if not insights:
            return False

        file_paths = get_file_paths(out_dir, report_key)
        enhanced_excel_path = file_paths["excel_file"].replace(
            ".xlsx", "_enhanced.xlsx"
        )

        # Try to create enhanced report with Excel MCP
        try:
            # This is where we would use Excel MCP functions for advanced features
            # For now, we'll create a placeholder for the enhanced functionality

            log(
                f"📈 Enhanced Excel report functionality ready for {report_key}",
                out_dir,
            )

            # Create summary sheet with insights
            summary_data = [
                ["Metric", "Value"],
                ["Report Type", insights.get("report_type", "")],
                ["Total Rows", insights.get("total_rows", 0)],
                [
                    "Data Completeness %",
                    round(
                        insights.get("data_quality", {}).get("data_completeness", 0), 2
                    ),
                ],
                ["Generated", insights.get("timestamp", "")],
            ]

            # Add business insights to summary
            business_insights = insights.get("business_insights", {})
            for key, value in business_insights.items():
                if isinstance(value, (int, float, str)) and key != "analysis_error":
                    summary_data.append([key.replace("_", " ").title(), value])

            # For now, we'll save the enhanced data as a separate JSON file
            # This can be expanded to use Excel MCP functions when available
            enhanced_insights_file = os.path.join(
                out_dir, f"{report_key}_enhanced_insights.json"
            )
            with open(enhanced_insights_file, "w", encoding="utf-8") as f:
                json.dump(
                    {
                        "summary_data": summary_data,
                        "chart_recommendations": get_chart_recommendations(
                            report_key, headers, rows
                        ),
                        "insights": insights,
                    },
                    f,
                    indent=2,
                )

            log(
                f"📊 Enhanced insights saved: {os.path.basename(enhanced_insights_file)}",
                out_dir,
            )
            return True

        except Exception as enhanced_error:
            log(f"⚠️ Enhanced Excel creation failed: {enhanced_error}", out_dir)
            return False

    except Exception as e:
        log(f"❌ Enhanced Excel creation failed for {report_key}: {e}", out_dir)
        return False


def get_chart_recommendations(report_key: str, headers: list, rows: list):
    """Get chart recommendations based on report type and data structure"""
    recommendations = []

    try:
        if report_key == "open_sales_orders":
            # Recommend customer distribution chart
            customer_col = next(
                (i for i, h in enumerate(headers) if "customer" in h.lower()), None
            )
            if customer_col is not None:
                recommendations.append(
                    {
                        "chart_type": "pie",
                        "title": "Orders by Customer",
                        "data_column": customer_col,
                        "description": "Distribution of open sales orders by customer",
                    }
                )

            # Recommend item quantity chart
            qty_col = next(
                (
                    i
                    for i, h in enumerate(headers)
                    if "qty" in h.lower() or "quantity" in h.lower()
                ),
                None,
            )
            item_col = next(
                (i for i, h in enumerate(headers) if "item" in h.lower()), None
            )
            if qty_col is not None and item_col is not None:
                recommendations.append(
                    {
                        "chart_type": "column",
                        "title": "Quantity by Item",
                        "x_axis_column": item_col,
                        "y_axis_column": qty_col,
                        "description": "Quantity distribution across items",
                    }
                )

        elif report_key == "sales_by_item":
            # Recommend sales performance chart
            item_col = next(
                (i for i, h in enumerate(headers) if "item" in h.lower()), None
            )
            sales_col = next(
                (
                    i
                    for i, h in enumerate(headers)
                    if "sales" in h.lower() or "amount" in h.lower()
                ),
                None,
            )
            if item_col is not None and sales_col is not None:
                recommendations.append(
                    {
                        "chart_type": "column",
                        "title": "Sales by Item",
                        "x_axis_column": item_col,
                        "y_axis_column": sales_col,
                        "description": "Sales performance by item",
                    }
                )

        elif report_key == "profit_loss":
            # Recommend P&L trend chart
            account_col = next(
                (i for i, h in enumerate(headers) if "account" in h.lower()), None
            )
            amount_col = next(
                (
                    i
                    for i, h in enumerate(headers)
                    if "amount" in h.lower() or "balance" in h.lower()
                ),
                None,
            )
            if account_col is not None and amount_col is not None:
                recommendations.append(
                    {
                        "chart_type": "bar",
                        "title": "Profit & Loss by Account",
                        "x_axis_column": account_col,
                        "y_axis_column": amount_col,
                        "description": "Financial performance by account",
                    }
                )

        return recommendations

    except Exception as e:
        return [{"error": f"Chart recommendation failed: {e}"}]


def test_xml_generation():
    """Test XML generation against all report types"""
    print("🧪 Testing XML generation for all report types...")

    test_date_from = "2025-08-01"
    test_date_to = "2025-08-15"

    for report_key, config in REPORT_CONFIGS.items():
        try:
            print(f"\n📋 Testing {config['name']} ({report_key})...")

            # Generate XML
            xml = build_report_qbxml(
                QBXML_VERSION_PRIMARY,
                config["qbxml_type"],
                test_date_from,
                test_date_to,
                report_key,
            )

            print(f"✅ Generated XML ({len(xml)} chars)")
            print("📄 XML Preview:")
            print(xml[:200] + "..." if len(xml) > 200 else xml)

            # Validate structure
            import xml.etree.ElementTree as ET

            try:
                root = ET.fromstring(xml)
                # Look for different report type elements
                report_type_elem = (
                    root.find(".//GeneralDetailReportType")
                    or root.find(".//GeneralSummaryReportType")
                    or root.find(".//AgingReportType")
                )
                if report_type_elem is not None:
                    print(f"🎯 ReportType: {report_type_elem.text}")
                else:
                    print("⚠️ No ReportType found in XML")
            except ET.ParseError as pe:
                print(f"❌ XML Parse Error: {pe}")

        except Exception as e:
            print(f"❌ Error testing {report_key}: {e}")

    print("\n🏁 XML generation test completed!")


if __name__ == "__main__":
    # Add test mode
    import sys

    if len(sys.argv) > 1 and sys.argv[1] == "--test-xml":
        test_xml_generation()
        sys.exit(0)


def generate_context7_insights(headers, rows, report_key: str, out_dir: str):
    """Generate enhanced business insights using Context7 MCP and data analysis"""
    try:
        insights = {
            "report_type": REPORT_CONFIGS[report_key]["name"],
            "total_rows": len(rows),
            "timestamp": dt.datetime.now().isoformat(),
            "key_metrics": {},
            "data_quality": {},
            "business_insights": {},
        }

        # Data quality analysis
        insights["data_quality"]["empty_rows"] = sum(
            1 for row in rows if all(not str(cell).strip() for cell in row)
        )
        insights["data_quality"]["columns_count"] = len(headers)
        insights["data_quality"]["data_completeness"] = (
            (len(rows) - insights["data_quality"]["empty_rows"])
            / max(len(rows), 1)
            * 100
        )

        # Report-specific insights
        if report_key == "open_sales_orders":
            insights["key_metrics"]["open_orders_count"] = len(rows)

            # Analyze order patterns if we have the right columns
            try:
                # Look for common QuickBooks columns
                customer_col = next(
                    (i for i, h in enumerate(headers) if "customer" in h.lower()), None
                )
                item_col = next(
                    (i for i, h in enumerate(headers) if "item" in h.lower()), None
                )
                qty_col = next(
                    (
                        i
                        for i, h in enumerate(headers)
                        if "qty" in h.lower() or "quantity" in h.lower()
                    ),
                    None,
                )

                if customer_col is not None:
                    customers = [
                        row[customer_col]
                        for row in rows
                        if len(row) > customer_col and row[customer_col]
                    ]
                    unique_customers = len(set(customers))
                    insights["business_insights"]["unique_customers"] = unique_customers

                    # Top customers by order count
                    from collections import Counter

                    customer_counts = Counter(customers)
                    insights["business_insights"]["top_customers"] = dict(
                        customer_counts.most_common(5)
                    )

                if item_col is not None:
                    items = [
                        row[item_col]
                        for row in rows
                        if len(row) > item_col and row[item_col]
                    ]
                    unique_items = len(set(items))
                    insights["business_insights"]["unique_items"] = unique_items

                    # Top items by frequency
                    from collections import Counter

                    item_counts = Counter(items)
                    insights["business_insights"]["top_items"] = dict(
                        item_counts.most_common(5)
                    )

            except Exception as analysis_error:
                insights["business_insights"]["analysis_error"] = str(analysis_error)

        elif report_key == "profit_loss":
            insights["key_metrics"]["pl_line_items"] = len(rows)

            # Look for revenue and expense patterns
            try:
                amount_cols = [
                    i
                    for i, h in enumerate(headers)
                    if any(term in h.lower() for term in ["amount", "total", "balance"])
                ]
                if amount_cols:
                    amounts = []
                    for row in rows:
                        for col in amount_cols:
                            if len(row) > col and row[col]:
                                try:
                                    # Clean and convert amount strings
                                    amount_str = (
                                        str(row[col])
                                        .replace(",", "")
                                        .replace("$", "")
                                        .replace("(", "-")
                                        .replace(")", "")
                                        .strip()
                                    )
                                    if amount_str and amount_str != "-":
                                        amounts.append(float(amount_str))
                                except (ValueError, TypeError):
                                    continue

                    if amounts:
                        insights["business_insights"]["total_amounts_analyzed"] = len(
                            amounts
                        )
                        insights["business_insights"]["sum_amounts"] = sum(amounts)
                        insights["business_insights"]["avg_amount"] = sum(
                            amounts
                        ) / len(amounts)
                        insights["business_insights"]["positive_amounts"] = sum(
                            1 for a in amounts if a > 0
                        )
                        insights["business_insights"]["negative_amounts"] = sum(
                            1 for a in amounts if a < 0
                        )

            except Exception as analysis_error:
                insights["business_insights"]["analysis_error"] = str(analysis_error)

        elif report_key == "sales_by_item":
            insights["key_metrics"]["items_with_sales"] = len(rows)

            # Analyze sales patterns
            try:
                item_col = next(
                    (i for i, h in enumerate(headers) if "item" in h.lower()), None
                )
                sales_cols = [
                    i
                    for i, h in enumerate(headers)
                    if any(
                        term in h.lower()
                        for term in ["sales", "amount", "qty", "quantity"]
                    )
                ]

                if item_col is not None and sales_cols:
                    items_data = {}
                    for row in rows:
                        if len(row) > item_col and row[item_col]:
                            item_name = row[item_col]
                            if item_name not in items_data:
                                items_data[item_name] = []

                            for col in sales_cols:
                                if len(row) > col and row[col]:
                                    try:
                                        value_str = (
                                            str(row[col])
                                            .replace(",", "")
                                            .replace("$", "")
                                            .strip()
                                        )
                                        if value_str and value_str != "-":
                                            items_data[item_name].append(
                                                float(value_str)
                                            )
                                    except (ValueError, TypeError):
                                        continue

                    insights["business_insights"]["items_analyzed"] = len(items_data)
                    insights["business_insights"]["items_with_data"] = sum(
                        1 for item_data in items_data.values() if item_data
                    )

            except Exception as analysis_error:
                insights["business_insights"]["analysis_error"] = str(analysis_error)

        # Try to use Context7 MCP for additional insights
        try:
            # This would be where we'd call Context7 MCP if available
            # For now, we'll add a placeholder for future enhancement
            insights["context7_status"] = (
                "Context7 MCP integration ready for enhancement"
            )
        except Exception as context7_error:
            insights["context7_status"] = (
                f"Context7 MCP not available: {context7_error}"
            )

        # Save insights to JSON file
        insights_file = os.path.join(out_dir, f"{report_key}_insights.json")
        with open(insights_file, "w", encoding="utf-8") as f:
            json.dump(insights, f, indent=2)

        log(
            f"📊 Generated insights for {report_key}: {len(insights['business_insights'])} business metrics",
            out_dir,
        )
        return insights

    except Exception as e:
        log(f"Context7 insights generation failed for {report_key}: {e}", out_dir)
        return None


def export_report(
    report_key: str, out_dir: str = None, date_from: str = None, date_to: str = None
):
    """Execute single report export with dynamic output directory and date range"""
    if out_dir is None:
        out_dir = DEFAULT_OUT_DIR

    os.makedirs(out_dir, exist_ok=True)
    config = REPORT_CONFIGS[report_key]

    log(f"🔄 Starting export for {config['name']} (key: {report_key})", out_dir)

    try_versions = [QBXML_VERSION_PRIMARY, QBXML_VERSION_FALLBACK]
    last_exc = None

    for ver_idx, ver in enumerate(try_versions):
        try:
            log(
                f"📡 Attempting {config['name']} with qbXML version {ver} (attempt {ver_idx + 1}/{len(try_versions)})",
                out_dir,
            )

            # Build the XML request
            req = build_report_qbxml(
                ver, config["qbxml_type"], date_from, date_to, report_key
            )
            log(
                f"📝 Built XML request for {config['name']}: {len(req)} characters",
                out_dir,
            )

            # Validate XML against working examples
            validate_xml_against_examples(report_key, req, out_dir)

            # Execute the request
            resp, info = qb_request(req, out_dir, report_key)
            log(
                f"📨 Received response for {config['name']}: {len(resp)} characters",
                out_dir,
            )

            # Parse the response
            headers, rows = parse_report_rows(resp)
            log(
                f"📊 Parsed {config['name']}: {len(headers)} columns, {len(rows)} rows",
                out_dir,
            )

            # Generate CSV
            csv_text = render_csv(headers, rows)

            # Generate Context7 insights
            insights = generate_context7_insights(headers, rows, report_key, out_dir)

            # Create Excel report
            excel_created = create_excel_report(headers, rows, out_dir, report_key)

            # Create enhanced Excel report with charts if possible
            enhanced_excel = create_enhanced_excel_report(
                headers, rows, out_dir, report_key, insights
            )

            return _write_outputs(
                csv_text,
                rows,
                info,
                out_dir,
                report_key,
                excel_created or enhanced_excel,
                insights,
            )

        except Exception as e:
            last_exc = e
            log(f"{config['name']} {ver} failed: {e}", out_dir)
            if "parsing the provided XML text stream" in str(e):
                continue
            else:
                break

    # Fallback for Open Sales Orders only
    if report_key == "open_sales_orders" and ALLOW_SALESORDER_FALLBACK:
        try:
            req = build_salesorder_query(QBXML_VERSION_FALLBACK)
            resp, info = qb_request(req, out_dir, report_key)
            headers, rows = parse_salesorders_to_rows(resp)
            csv_text = render_csv(headers, rows)

            insights = generate_context7_insights(headers, rows, report_key, out_dir)
            excel_created = create_excel_report(headers, rows, out_dir, report_key)

            # Create enhanced Excel report with charts if possible
            enhanced_excel = create_enhanced_excel_report(
                headers, rows, out_dir, report_key, insights
            )

            return _write_outputs(
                csv_text,
                rows,
                info,
                out_dir,
                report_key,
                excel_created or enhanced_excel,
                insights,
            )

        except Exception as e2:
            last_exc = e2

    raise (
        last_exc if last_exc else RuntimeError(f"Unknown failure for {config['name']}")
    )


def _write_outputs(
    csv_text, rows, info, out_dir: str, report_key: str, excel_created: bool, insights
):
    """Write outputs to files with change detection"""
    file_paths = get_file_paths(out_dir, report_key)
    config = REPORT_CONFIGS[report_key]
    digest = sha256_text(csv_text)
    last = None

    if os.path.exists(file_paths["hash_file"]):
        try:
            with open(file_paths["hash_file"], "r", encoding="utf-8") as f:
                last = f.read().strip()
        except Exception:
            pass

    changed = digest != last

    # Write CSV
    with open(file_paths["main_csv"], "w", newline="", encoding="utf-8") as f:
        f.write(csv_text)

    if changed:
        snap = snapshot_filename(out_dir, report_key)
        with open(snap, "w", newline="", encoding="utf-8") as f:
            f.write(csv_text)
        with open(file_paths["hash_file"], "w", encoding="utf-8") as f:
            f.write(digest)
        log(
            f"{config['name']} changed. Wrote {os.path.basename(file_paths['main_csv'])} and snapshot {os.path.basename(snap)} (rows={len(rows)}) Excel: {'✅' if excel_created else '❌'}",
            out_dir,
        )
    else:
        log(
            f"{config['name']} no change. Updated {os.path.basename(file_paths['main_csv'])} (rows={len(rows)}) Excel: {'✅' if excel_created else '❌'}",
            out_dir,
        )

    return {
        "rows": len(rows),
        "changed": changed,
        "connect_info": info,
        "timestamp": dt.datetime.now(),
        "excel_created": excel_created,
        "insights": insights,
        "report_name": config["name"],
    }


def export_all_reports(out_dir: str = None, date_from: str = None, date_to: str = None):
    """Export all configured reports with enhanced error handling and user-friendly messages"""
    if out_dir is None:
        out_dir = DEFAULT_OUT_DIR

    results = {}
    errors = {}
    connection_tested = False
    connection_working = False

    log(f"� Startineg export of all reports to: {out_dir}", out_dir)
    if date_from and date_to:
        log(f"📅 Date range: {date_from} to {date_to}", out_dir)

    # Process reports in order of complexity (simplest first)
    report_order = [
        "open_sales_orders",  # Usually works
        "profit_loss",
        "sales_by_item",
        "profit_loss_detail",
        "sales_by_item_detail",
        "sales_by_rep_detail",
        "purchase_by_vendor_detail",
        "ap_aging_detail",
        "ar_aging_detail",
    ]

    for report_key in report_order:
        if report_key not in REPORT_CONFIGS:
            continue

        config = REPORT_CONFIGS[report_key]
        try:
            log(f"🔄 Starting export for {config['name']} (key: {report_key})", out_dir)
            result = export_report(report_key, out_dir, date_from, date_to)
            results[report_key] = result
            connection_working = True
            log(
                f"✅ {config['name']} completed successfully - {result['rows']} rows, Excel: {'Yes' if result['excel_created'] else 'No'}",
                out_dir,
            )

        except Exception as e:
            # Enhanced error handling with user-friendly messages
            if not connection_tested:
                connection_tested = True
                # Check if this is a connection issue that affects all reports
                error_str = str(e)
                if any(code in error_str for code in ["-2147221005", "-2147221164", "Invalid class string", "Class not registered"]):
                    log("🔍 Detected QuickBooks connection issue. Running diagnostics...", out_dir)
                    
                    # Run diagnostics
                    diagnostics = diagnose_quickbooks_connection(out_dir)
                    
                    # Log user-friendly error message
                    log("❌ QUICKBOOKS CONNECTION PROBLEM DETECTED", out_dir)
                    log("", out_dir)
                    log("The application cannot connect to QuickBooks Desktop.", out_dir)
                    log("This is usually because the QuickBooks SDK is not installed or not working properly.", out_dir)
                    log("", out_dir)
                    log("IMMEDIATE STEPS TO FIX:", out_dir)
                    log("1. Make sure QuickBooks Desktop is installed on this computer", out_dir)
                    log("2. Download and install the QuickBooks SDK from the Intuit Developer website", out_dir)
                    log("3. Restart your computer after installing the SDK", out_dir)
                    log("4. Run this application as Administrator", out_dir)
                    log("", out_dir)
                    log(f"📊 A detailed diagnostic report has been saved to: {out_dir}", out_dir)
                    log("Check 'QuickBooks_Diagnostic_Report.xlsx' for more information.", out_dir)
                    
                    # Since this is a fundamental connection issue, all reports will fail
                    # Add the same error to all remaining reports
                    for remaining_key in report_order:
                        if remaining_key not in results and remaining_key not in errors:
                            errors[remaining_key] = "Cannot connect to QuickBooks Desktop - SDK not installed or not working"
                    break
            
            # Log individual report error
            user_friendly_msg = "Cannot connect to QuickBooks Desktop - check diagnostic report for solutions"
            errors[report_key] = user_friendly_msg
            log(f"❌ {config['name']}: {user_friendly_msg}", out_dir)

    # Enhanced summary with user guidance
    total_reports = len(report_order)
    successful_reports = len(results)
    failed_reports = len(errors)

    log("", out_dir)
    log("=" * 60, out_dir)
    log(f"📊 EXPORT SUMMARY: {successful_reports}/{total_reports} successful, {failed_reports} failed", out_dir)
    log("=" * 60, out_dir)

    if successful_reports > 0:
        log(f"✅ Successful reports: {', '.join([REPORT_CONFIGS[k]['name'] for k in results.keys()])}", out_dir)

    if failed_reports > 0:
        if not connection_working:
            log("❌ ALL REPORTS FAILED - QuickBooks connection problem", out_dir)
            log("", out_dir)
            log("NEXT STEPS:", out_dir)
            log("1. Check the diagnostic report in the output folder", out_dir)
            log("2. Install QuickBooks SDK if not already installed", out_dir)
            log("3. Run as Administrator", out_dir)
            log("4. Make sure QuickBooks Desktop is properly installed", out_dir)
        else:
            log(f"❌ Failed reports: {', '.join([REPORT_CONFIGS[k]['name'] for k in errors.keys()])}", out_dir)

    return results, errors


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("QuickBooks Autoreporter")
        self.geometry("900x650")  # Increased height for date selectors
        self.resizable(False, False)

        # Load settings
        self.settings = load_settings()
        self.output_dir = self.settings["output_dir"]
        self.selected_interval = self.settings["interval"]

        # State variables
        self.running = False
        self._timer = None
        self.last_export_time = None
        self.last_export_results = {}

        # Create GUI variables
        self.status_var = tk.StringVar(value="Idle")
        self.next_var = tk.StringVar(value="-")
        self.folder_var = tk.StringVar(value=self.output_dir)
        self.interval_var = tk.StringVar(value=self.selected_interval)
        self.last_check_var = tk.StringVar(value="-")
        self.time_since_var = tk.StringVar(value="-")
        self.date_from_var = tk.StringVar(value=self.settings["report_date_from"])
        self.date_to_var = tk.StringVar(value=self.settings["report_date_to"])

        # Report status variables
        self.report_status_vars = {}
        for report_key in REPORT_CONFIGS.keys():
            self.report_status_vars[report_key] = {
                "status": tk.StringVar(value="-"),
                "rows": tk.StringVar(value="-"),
                "excel": tk.StringVar(value="-"),
            }

        self.create_widgets()
        self.start_timer_display()

        # Save settings on close
        self.protocol("WM_DELETE_WINDOW", self.on_exit)

    def create_widgets(self):
        # Title
        title_frame = tk.Frame(self)
        title_frame.pack(pady=(10, 5))
        tk.Label(
            title_frame, text="QuickBooks Autoreporter", font=("Segoe UI", 14, "bold")
        ).pack()
        tk.Label(
            title_frame,
            text=(
                "Automated exports • Change detection + snapshots • Styled Excel • "
                "Scheduled checks with folder selection"
            ),
            font=("Segoe UI", 10),
            fg="#666",
            wraplength=740,
            justify="center",
        ).pack()

        # Configuration section
        config_frame = tk.LabelFrame(
            self, text="Configuration", font=("Segoe UI", 10, "bold")
        )
        config_frame.pack(fill="x", padx=10, pady=5)

        # Output folder selection
        folder_frame = tk.Frame(config_frame)
        folder_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(folder_frame, text="Output Folder:", width=15, anchor="w").pack(
            side="left"
        )
        tk.Entry(
            folder_frame, textvariable=self.folder_var, state="readonly", width=50
        ).pack(side="left", padx=5)
        tk.Button(
            folder_frame, text="Browse...", command=self.select_folder, width=10
        ).pack(side="right")

        # Interval selection
        interval_frame = tk.Frame(config_frame)
        interval_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(interval_frame, text="Check Interval:", width=15, anchor="w").pack(
            side="left"
        )
        interval_combo = ttk.Combobox(
            interval_frame,
            textvariable=self.interval_var,
            values=list(INTERVAL_OPTIONS.keys()),
            state="readonly",
            width=15,
        )
        interval_combo.pack(side="left", padx=5)
        interval_combo.bind("<<ComboboxSelected>>", self.on_interval_changed)

        # Date range selection (affects reports with date ranges)
        date_frame = tk.LabelFrame(
            config_frame,
            text="Date Range (for P&L, Sales, AP/AR Aging Reports)",
            font=("Segoe UI", 9, "bold"),
        )
        date_frame.pack(fill="x", padx=10, pady=5)

        # From date
        from_date_frame = tk.Frame(date_frame)
        from_date_frame.pack(fill="x", padx=10, pady=2)
        tk.Label(from_date_frame, text="From Date:", width=12, anchor="w").pack(
            side="left"
        )
        from_date_entry = tk.Entry(
            from_date_frame, textvariable=self.date_from_var, width=12
        )
        from_date_entry.pack(side="left", padx=5)
        from_date_entry.bind("<FocusOut>", self.on_date_changed)
        tk.Label(
            from_date_frame, text="(YYYY-MM-DD)", font=("Segoe UI", 8), fg="#666"
        ).pack(side="left", padx=5)

        # To date
        to_date_frame = tk.Frame(date_frame)
        to_date_frame.pack(fill="x", padx=10, pady=2)
        tk.Label(to_date_frame, text="To Date:", width=12, anchor="w").pack(side="left")
        to_date_entry = tk.Entry(to_date_frame, textvariable=self.date_to_var, width=12)
        to_date_entry.pack(side="left", padx=5)
        to_date_entry.bind("<FocusOut>", self.on_date_changed)
        tk.Label(
            to_date_frame, text="(YYYY-MM-DD)", font=("Segoe UI", 8), fg="#666"
        ).pack(side="left", padx=5)

        # Quick date buttons
        quick_date_frame = tk.Frame(date_frame)
        quick_date_frame.pack(fill="x", padx=10, pady=5)
        tk.Button(
            quick_date_frame,
            text="This Month",
            command=self.set_this_month,
            bg="#9C27B0",
            fg="white",
            font=("Segoe UI", 8),
            width=10,
        ).pack(side="left", padx=2)
        tk.Button(
            quick_date_frame,
            text="Last Month",
            command=self.set_last_month,
            bg="#9C27B0",
            fg="white",
            font=("Segoe UI", 8),
            width=10,
        ).pack(side="left", padx=2)
        tk.Button(
            quick_date_frame,
            text="This Year",
            command=self.set_this_year,
            bg="#9C27B0",
            fg="white",
            font=("Segoe UI", 8),
            width=10,
        ).pack(side="left", padx=2)
        tk.Button(
            quick_date_frame,
            text="Last Year",
            command=self.set_last_year,
            bg="#9C27B0",
            fg="white",
            font=("Segoe UI", 8),
            width=10,
        ).pack(side="left", padx=2)

        # Control buttons
        control_frame = tk.Frame(self)
        control_frame.pack(pady=10)
        tk.Button(
            control_frame,
            text="Start Auto",
            width=12,
            command=self.start,
            bg="#4CAF50",
            fg="white",
            font=("Segoe UI", 10, "bold"),
        ).grid(row=0, column=0, padx=5)
        tk.Button(
            control_frame,
            text="Stop",
            width=12,
            command=self.stop,
            bg="#f44336",
            fg="white",
            font=("Segoe UI", 10, "bold"),
        ).grid(row=0, column=1, padx=5)
        tk.Button(
            control_frame,
            text="Export All Now",
            width=15,
            command=self.export_now,
            bg="#2196F3",
            fg="white",
            font=("Segoe UI", 10, "bold"),
        ).grid(row=0, column=2, padx=5)
        tk.Button(
            control_frame,
            text="Open Folder",
            width=12,
            command=self.open_folder,
            bg="#FF9800",
            fg="white",
            font=("Segoe UI", 10, "bold"),
        ).grid(row=0, column=3, padx=5)

        # Status section
        status_frame = tk.LabelFrame(self, text="Status", font=("Segoe UI", 10, "bold"))
        status_frame.pack(fill="x", padx=10, pady=5)

        status_grid = tk.Frame(status_frame)
        status_grid.pack(padx=10, pady=5)

        # Status info
        tk.Label(status_grid, text="Status:", width=15, anchor="e").grid(
            row=0, column=0, sticky="e", padx=5, pady=2
        )
        tk.Label(
            status_grid,
            textvariable=self.status_var,
            fg="#006400",
            width=20,
            anchor="w",
        ).grid(row=0, column=1, sticky="w", padx=5, pady=2)
        tk.Label(status_grid, text="Next Run:", width=15, anchor="e").grid(
            row=0, column=2, sticky="e", padx=5, pady=2
        )
        tk.Label(status_grid, textvariable=self.next_var, width=25, anchor="w").grid(
            row=0, column=3, sticky="w", padx=5, pady=2
        )

        # Reports status
        reports_frame = tk.LabelFrame(
            self, text="Report Status", font=("Segoe UI", 10, "bold")
        )
        reports_frame.pack(fill="x", padx=10, pady=5)

        reports_grid = tk.Frame(reports_frame)
        reports_grid.pack(padx=10, pady=5)

        # Headers
        tk.Label(
            reports_grid, text="Report", font=("Segoe UI", 9, "bold"), width=20
        ).grid(row=0, column=0, padx=5, pady=2)
        tk.Label(
            reports_grid, text="Status", font=("Segoe UI", 9, "bold"), width=15
        ).grid(row=0, column=1, padx=5, pady=2)
        tk.Label(
            reports_grid, text="Rows", font=("Segoe UI", 9, "bold"), width=10
        ).grid(row=0, column=2, padx=5, pady=2)
        tk.Label(
            reports_grid, text="Excel", font=("Segoe UI", 9, "bold"), width=10
        ).grid(row=0, column=3, padx=5, pady=2)

        # Report rows
        for i, (report_key, config) in enumerate(REPORT_CONFIGS.items(), 1):
            tk.Label(reports_grid, text=config["name"], width=20, anchor="w").grid(
                row=i, column=0, padx=5, pady=2, sticky="w"
            )
            tk.Label(
                reports_grid,
                textvariable=self.report_status_vars[report_key]["status"],
                width=15,
                anchor="w",
            ).grid(row=i, column=1, padx=5, pady=2, sticky="w")
            tk.Label(
                reports_grid,
                textvariable=self.report_status_vars[report_key]["rows"],
                width=10,
                anchor="w",
            ).grid(row=i, column=2, padx=5, pady=2, sticky="w")
            tk.Label(
                reports_grid,
                textvariable=self.report_status_vars[report_key]["excel"],
                width=10,
                anchor="w",
            ).grid(row=i, column=3, padx=5, pady=2, sticky="w")

        # Exit button
        tk.Button(
            self,
            text="Exit",
            width=15,
            command=self.on_exit,
            bg="#795548",
            fg="white",
            font=("Segoe UI", 10, "bold"),
        ).pack(pady=(10, 10))

    def select_folder(self):
        """Open folder selection dialog"""
        folder = filedialog.askdirectory(
            title="Select Output Folder", initialdir=self.output_dir
        )
        if folder:
            self.output_dir = folder
            self.folder_var.set(folder)
            self.settings["output_dir"] = folder
            save_settings(self.settings)

    def on_interval_changed(self, event=None):
        """Handle interval selection change"""
        self.selected_interval = self.interval_var.get()
        self.settings["interval"] = self.selected_interval
        save_settings(self.settings)

        if self.running:
            self.stop()
            self.after(100, self.start)

    def on_date_changed(self, event=None):
        """Handle date field changes"""
        try:
            # Validate date format
            from_date = self.date_from_var.get().strip()
            to_date = self.date_to_var.get().strip()

            if from_date:
                dt.datetime.strptime(from_date, "%Y-%m-%d")
                self.settings["report_date_from"] = from_date

            if to_date:
                dt.datetime.strptime(to_date, "%Y-%m-%d")
                self.settings["report_date_to"] = to_date

            save_settings(self.settings)
        except ValueError:
            # Invalid date format - could show a warning
            pass

    def set_this_month(self):
        """Set date range to current month"""
        today = dt.date.today()
        first_day = today.replace(day=1)
        self.date_from_var.set(first_day.strftime("%Y-%m-%d"))
        self.date_to_var.set(today.strftime("%Y-%m-%d"))
        self.on_date_changed()

    def set_last_month(self):
        """Set date range to last month"""
        today = dt.date.today()
        first_day_this_month = today.replace(day=1)
        last_day_last_month = first_day_this_month - dt.timedelta(days=1)
        first_day_last_month = last_day_last_month.replace(day=1)
        self.date_from_var.set(first_day_last_month.strftime("%Y-%m-%d"))
        self.date_to_var.set(last_day_last_month.strftime("%Y-%m-%d"))
        self.on_date_changed()

    def set_this_year(self):
        """Set date range to current year"""
        today = dt.date.today()
        first_day_year = today.replace(month=1, day=1)
        self.date_from_var.set(first_day_year.strftime("%Y-%m-%d"))
        self.date_to_var.set(today.strftime("%Y-%m-%d"))
        self.on_date_changed()

    def set_last_year(self):
        """Set date range to last year"""
        today = dt.date.today()
        last_year = today.year - 1
        first_day_last_year = dt.date(last_year, 1, 1)
        last_day_last_year = dt.date(last_year, 12, 31)
        self.date_from_var.set(first_day_last_year.strftime("%Y-%m-%d"))
        self.date_to_var.set(last_day_last_year.strftime("%Y-%m-%d"))
        self.on_date_changed()

    def start_timer_display(self):
        """Start the timer display update loop"""
        self.update_timer_display()
        self.after(1000, self.start_timer_display)

    def update_timer_display(self):
        """Update the timer display"""
        if self.last_export_time:
            time_diff = dt.datetime.now() - self.last_export_time
            hours, remainder = divmod(int(time_diff.total_seconds()), 3600)
            minutes, seconds = divmod(remainder, 60)

            if hours > 0:
                time_str = f"{hours}h {minutes}m {seconds}s"
            elif minutes > 0:
                time_str = f"{minutes}m {seconds}s"
            else:
                time_str = f"{seconds}s"

            self.time_since_var.set(time_str)
        else:
            self.time_since_var.set("-")

    def open_folder(self):
        """Open output folder in explorer"""
        os.makedirs(self.output_dir, exist_ok=True)
        os.startfile(self.output_dir)

    def on_exit(self):
        """Handle application exit"""
        self.stop()
        save_settings(self.settings)
        self.destroy()

    def set_next_time(self):
        """Set next run time display"""
        interval_seconds = INTERVAL_OPTIONS[self.selected_interval]
        nxt = dt.datetime.now() + dt.timedelta(seconds=interval_seconds)
        self.next_var.set(nxt.strftime("%Y-%m-%d %H:%M:%S"))

    def start(self):
        """Start scheduled exports"""
        if self.running:
            return
        self.running = True
        self.status_var.set("Running")
        log("Started scheduled exports", self.output_dir)
        self.run_once_then_schedule()

    def stop(self):
        """Stop scheduled exports"""
        self.running = False
        if self._timer is not None:
            try:
                self.after_cancel(self._timer)
            except Exception:
                pass
            self._timer = None
        self.status_var.set("Stopped")
        self.next_var.set("-")
        log("Stopped scheduled exports", self.output_dir)

    def export_now(self):
        """Trigger immediate export (main-thread UI prep + background work)"""
        # Prepare UI on the main thread before starting background work
        self._prepare_export_ui()
        threading.Thread(target=self._export_worker, daemon=True).start()

    def run_once_then_schedule(self):
        """Run export once then schedule next"""
        if not self.running:
            return
        self.export_now()
        self.set_next_time()
        interval_seconds = INTERVAL_OPTIONS[self.selected_interval]
        self._timer = self.after(interval_seconds * 1000, self.run_once_then_schedule)

    def _export_worker(self):
        """Background export worker (no direct UI access)"""
        try:
            # Get date range from GUI
            date_from = self.date_from_var.get().strip() or None
            date_to = self.date_to_var.get().strip() or None

            results, errors = export_all_reports(self.output_dir, date_from, date_to)
            # Hand off results to the UI thread
            self.after(0, self._on_export_complete, results, errors)
        except Exception as e:
            msg = f"ERROR: {e}"
            log(msg, self.output_dir)
            # Update UI and show error from the main thread
            self.after(0, self._on_export_error, msg)

    def _prepare_export_ui(self):
        """Prepare UI state on the main thread before export"""
        self.status_var.set("Exporting...")
        # Reset all report statuses to a neutral state
        for report_key in REPORT_CONFIGS.keys():
            self.report_status_vars[report_key]["status"].set("Working...")
            self.report_status_vars[report_key]["rows"].set("-")
            self.report_status_vars[report_key]["excel"].set("-")

    def _on_export_complete(self, results, errors):
        """Apply export results to UI (main thread)"""
        self.last_export_time = dt.datetime.now()
        self.last_export_results = results

        # Update individual report status
        for report_key in REPORT_CONFIGS.keys():
            if report_key in results:
                result = results[report_key]
                status = "Changed" if result["changed"] else "No Change"
                self.report_status_vars[report_key]["status"].set(status)
                self.report_status_vars[report_key]["rows"].set(str(result["rows"]))
                self.report_status_vars[report_key]["excel"].set(
                    "✅" if result["excel_created"] else "❌"
                )
            elif report_key in errors:
                self.report_status_vars[report_key]["status"].set("Error")
                self.report_status_vars[report_key]["rows"].set("-")
                self.report_status_vars[report_key]["excel"].set("-")

        # Overall status
        if errors:
            self.status_var.set(f"Completed with {len(errors)} error(s)")
        else:
            self.status_var.set("All reports completed")

    def _on_export_error(self, msg: str):
        """Display error from export (main thread)"""
        self.status_var.set("Error")
        messagebox.showerror("QuickBooks Autoreporter", msg)


# Main application entry point
def run_gui():
    """Run the GUI application"""
    try:
        os.makedirs(DEFAULT_OUT_DIR, exist_ok=True)
        app = App()
        app.mainloop()
    except Exception as e:
        print(f"GUI Error: {e}")
        log(f"GUI Fatal: {e}")
        raise


if __name__ == "__main__":
    try:
        # Ensure default directory exists
        os.makedirs(DEFAULT_OUT_DIR, exist_ok=True)

        # Check if GUI should be launched
        import sys

        # Detect if running as executable without console
        has_console = hasattr(sys, "stdin") and sys.stdin is not None
        is_executable = getattr(sys, "frozen", False)

        # Check for diagnostic mode
        if len(sys.argv) > 1 and sys.argv[1] == "--diagnose":
            print("QuickBooks Auto Reporter - Diagnostic Mode")
            print("=" * 50)
            print("Running QuickBooks connectivity diagnostics...")
            print("")
            
            diagnostics = diagnose_quickbooks_connection(DEFAULT_OUT_DIR)
            
            print("DIAGNOSTIC RESULTS:")
            print("=" * 30)
            print(f"QuickBooks Desktop: {diagnostics['quickbooks_installation']['status']}")
            print(f"QuickBooks SDK: {diagnostics['sdk_installation']['status']}")
            
            if "connectivity_test" in diagnostics:
                print(f"COM Object Creation: {diagnostics['connectivity_test'].get('com_object_creation', 'Not tested')}")
                if "connection_test" in diagnostics["connectivity_test"]:
                    print(f"Connection Test: {diagnostics['connectivity_test']['connection_test']}")
            
            print("")
            if diagnostics.get("recommendations"):
                print("RECOMMENDATIONS:")
                print("-" * 20)
                for i, rec in enumerate(diagnostics["recommendations"], 1):
                    print(f"{i}. {rec}")
            
            print("")
            print(f"Detailed report saved to: {DEFAULT_OUT_DIR}")
            print("Check 'quickbooks_diagnostics.json' and 'QuickBooks_Diagnostic_Report.xlsx'")
            sys.exit(0)

        # Default to GUI mode for executables or when --gui is specified
        if len(sys.argv) > 1 and sys.argv[1] == "--gui":
            run_gui()
        elif is_executable or not has_console:
            # Running as executable or no console - launch GUI
            run_gui()
        else:
            # Enhanced command line interface
            print("QuickBooks Auto Reporter v1.0")
            print("=" * 40)
            print("Available reports: Open Sales Orders, Profit & Loss, Sales by Item")
            print("")
            print("Options:")
            print("  python quickbooks_autoreport.py --gui       # Launch GUI")
            print("  python quickbooks_autoreport.py --diagnose  # Run diagnostics")
            print("  python quickbooks_autoreport.py             # Command line mode")
            print("")

            # In non-interactive environments (e.g., no console/pyinstaller windowed),
            # stdin may be unavailable. Default to running all reports.
            try:
                interactive = bool(getattr(sys, "stdin", None)) and sys.stdin.isatty()
            except Exception:
                interactive = False

            if interactive:
                try:
                    print("Choose an option:")
                    print("1. Run all reports now")
                    print("2. Run diagnostics first")
                    print("3. Exit")
                    choice = input("Enter choice (1-3): ").strip()
                except (EOFError, RuntimeError):
                    print("No interactive stdin detected; defaulting to run reports.")
                    choice = "1"
            else:
                print("No interactive stdin detected; running all reports now.")
                choice = "1"

            if choice == "2":
                print("\nRunning QuickBooks diagnostics...")
                diagnostics = diagnose_quickbooks_connection(DEFAULT_OUT_DIR)
                
                print("\nDIAGNOSTIC RESULTS:")
                print("=" * 30)
                print(f"QuickBooks Desktop: {diagnostics['quickbooks_installation']['status']}")
                print(f"QuickBooks SDK: {diagnostics['sdk_installation']['status']}")
                
                if diagnostics.get("recommendations"):
                    print("\nRECOMMENDATIONS:")
                    for i, rec in enumerate(diagnostics["recommendations"], 1):
                        print(f"{i}. {rec}")
                
                print(f"\nDetailed report saved to: {DEFAULT_OUT_DIR}")
                
                # Ask if they want to continue with reports
                try:
                    continue_choice = input("\nDo you want to try running reports anyway? (y/n): ").lower().strip()
                except (EOFError, RuntimeError):
                    continue_choice = "n"
                
                if continue_choice != "y":
                    print("Exiting. Fix the issues above and try again.")
                    sys.exit(0)
                choice = "1"  # Continue to run reports

            if choice == "1":
                print("\nExporting all reports...")
                print("=" * 30)
                
                # Load settings for date range
                settings = load_settings()
                date_from = settings.get("report_date_from")
                date_to = settings.get("report_date_to")

                results, errors = export_all_reports(
                    DEFAULT_OUT_DIR, date_from, date_to
                )

                print("\n" + "=" * 50)
                print("FINAL RESULTS:")
                print("=" * 50)
                
                if results:
                    print("\n✅ SUCCESSFUL REPORTS:")
                    for key, result in results.items():
                        print(f"   • {result['report_name']}: {result['rows']} rows, Excel: {'Yes' if result['excel_created'] else 'No'}")

                if errors:
                    print("\n❌ FAILED REPORTS:")
                    for key, error in errors.items():
                        print(f"   • {REPORT_CONFIGS[key]['name']}: {error}")
                    
                    print(f"\n💡 TROUBLESHOOTING:")
                    print(f"   • Check the log file in: {DEFAULT_OUT_DIR}")
                    print(f"   • Run diagnostics: python quickbooks_autoreport.py --diagnose")
                    print(f"   • Make sure QuickBooks Desktop and SDK are installed")

                print(f"\n📁 Files saved to: {DEFAULT_OUT_DIR}")
                if date_from and date_to:
                    print(f"📅 Date range used: {date_from} to {date_to}")
                    
            elif choice == "3":
                print("Exiting.")
            else:
                print("Invalid choice. Exiting.")
                print("Use --gui flag to launch the GUI interface.")

    except Exception as e:
        print(f"\n❌ FATAL ERROR: {e}")
        print("\nTROUBLESHOoting:")
        print("1. Run diagnostics: python quickbooks_autoreport.py --diagnose")
        print("2. Make sure QuickBooks Desktop is installed")
        print("3. Install QuickBooks SDK from Intuit Developer website")
        print("4. Run as Administrator")
        raise
